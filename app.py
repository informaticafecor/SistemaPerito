"""
SistemaPerito - Sistema de Gestión de Asignaciones de Peritos
Autor: Desarrollado para control de asignaciones
Fecha: 2025
Descripción: Sistema web para gestionar asignaciones de peritos con validación
de disponibilidad, búsqueda avanzada y exportación de reportes.
"""

from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, session
import sqlite3
import json
from datetime import datetime, timedelta
from functools import wraps
import bcrypt
import os



from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from werkzeug.utils import secure_filename  



# Inicializar aplicación Flask
app = Flask(__name__)
app.config['JSON_AS_ASCII'] = False  # Para caracteres especiales en español

# Configuración para archivos
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx', 'jpg', 'jpeg', 'png'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16 MB máximo por archivo

# Crear carpeta si no existe
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    """
    Verifica si el archivo tiene una extensión permitida
    """
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# Configuración de sesiones
app.secret_key = 'tu_clave_secreta_muy_segura_2025'  # Cambiar en producción
app.config['SESSION_TYPE'] = 'filesystem'
app.config['PERMANENT_SESSION_LIFETIME'] = 28800  # 8 horas en segundos

# ============================================================================
# CONFIGURACIÓN DE BASE DE DATOS
# ============================================================================

def init_db():
    """
    Inicializa la base de datos SQLite creando las tablas necesarias
    si no existen. Se ejecuta al iniciar la aplicación.
    """
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    
    # Tabla de peritos con información básica
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS peritos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre_completo TEXT NOT NULL,
            tipo TEXT NOT NULL,
            estado TEXT DEFAULT 'Activo',
            fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Tabla de asignaciones con toda la información del Excel
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS asignaciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            hoja_envio TEXT,
            expediente TEXT,
            dependencia TEXT,
            tipo_perito TEXT,
            carpeta_fiscal TEXT,
            denominacion TEXT,  -- ← VERIFICA QUE EXISTA
            observaciones TEXT,
            lugar TEXT,
            fecha_inicio TEXT NOT NULL,
            fecha_fin TEXT NOT NULL,
            perito_asignado TEXT NOT NULL,
            perito_id INTEGER,
            desginacion TEXT,
            oficio_desplazamiento TEXT,
            estado TEXT DEFAULT 'Pendiente',
            fecha_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (perito_id) REFERENCES peritos (id)
        )
    ''')

    # ⬇️ AGREGAR ESTA TABLA ⬇️
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS archivos_asignaciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            asignacion_id INTEGER NOT NULL,
            nombre_original TEXT NOT NULL,
            nombre_guardado TEXT NOT NULL,
            tamano INTEGER,
            extension TEXT,
            fecha_subida TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            usuario_id INTEGER,
            usuario_nombre TEXT,
            FOREIGN KEY (asignacion_id) REFERENCES asignaciones(id),
            FOREIGN KEY (usuario_id) REFERENCES usuarios(id)
        )
    ''')
    print("✅ Tabla 'archivos_asignaciones' verificada/creada")
    # ⬆️ HASTA AQUÍ ⬆️


    
    # Tabla de historial para auditoría
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS historial (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            asignacion_id INTEGER,
            accion TEXT NOT NULL,
            detalles TEXT,
            fecha_hora TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (asignacion_id) REFERENCES asignaciones (id)
        )
    ''')
    
    conn.commit()
    
    # Insertar peritos iniciales si la tabla está vacía
    cursor.execute('SELECT COUNT(*) FROM peritos')
    if cursor.fetchone()[0] == 0:
        peritos_iniciales = [
            # Peritos Informáticos
            ('ALBERTO HONORATO BLACIDO QUITO', 'Informático'),
            ('MISAEL EDSON PALOMINO AYLAS', 'Informático'),
            ('LUIS ALBERTO VILLANUEVA HUAMAN', 'Informático'),
            # Peritos Acústicos
            ('MARCIAL SULCA CAHUANA', 'Acústico'),
            ('EDILBERTO EDISON ZAVALA CAMPOS', 'Acústico'),
            ('WILBER PAUL ESPINOZA LAUREANO', 'Acústico'),
            # Peritos Antropólogos
            ('BRIAN BARRY SOTO ALCAZAR', 'Antropólogo'),
            ('SANDRA LISBET IBARRA APAZA', 'Antropólogo'),
            # Peritos Contables
            ('ROSARIO CORDERO BORJA', 'Contable'),
            ('ANGELA ROXANA CALDERON BUSTAMANTE', 'Contable'),
            ('YESENIA ELIZABETH CHAVEZ VALERO', 'Contable')
        ]
        
        cursor.executemany(
            'INSERT INTO peritos (nombre_completo, tipo) VALUES (?, ?)',
            peritos_iniciales
        )
        conn.commit()


    # ============================================
    # TABLA DE VACACIONES ----------------------------------------------------------------------------------------
    # ============================================
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS vacaciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            perito_id INTEGER NOT NULL,
            tipo_perito TEXT,
            fecha_inicio TEXT NOT NULL,
            fecha_fin TEXT NOT NULL,
            dias_totales INTEGER,
            tipo_vacaciones TEXT DEFAULT 'Programadas',
            estado TEXT DEFAULT 'Aprobadas',
            observaciones TEXT,
            archivos_adjuntos TEXT,
            fecha_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (perito_id) REFERENCES peritos(id)
        )
    ''')
    
    print("✅ Tabla 'vacaciones' verificada/creada")
    conn.commit()


# ============================================
    # TABLA DE USUARIOS
    # ============================================
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            usuario TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            rol TEXT NOT NULL,
            perito_id INTEGER,
            nombre_completo TEXT,
            email TEXT,
            activo INTEGER DEFAULT 1,
            primer_inicio INTEGER DEFAULT 1,
            fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            ultimo_acceso TIMESTAMP,
            FOREIGN KEY (perito_id) REFERENCES peritos(id)
        )
    ''')
    print("✅ Tabla 'usuarios' verificada/creada")
    
    # ============================================
    # TABLA DE AUDITORÍA
    # ============================================
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS auditoria (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            usuario_id INTEGER NOT NULL,
            usuario_nombre TEXT,
            accion TEXT NOT NULL,
            modulo TEXT,
            descripcion TEXT,
            registro_id INTEGER,
            ip_address TEXT,
            fecha_hora TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (usuario_id) REFERENCES usuarios(id)
        )
    ''')
    print("✅ Tabla 'auditoria' verificada/creada")
    
    # ============================================
    # TABLA DE ACTIVIDADES DE PERITOS
    # ============================================
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS actividades_peritos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            perito_id INTEGER NOT NULL,
            tipo_perito TEXT NOT NULL,
            dependencia TEXT,
            carpeta_fiscal TEXT,
            denominacion TEXT,
            tipo_actividad TEXT NOT NULL,
            apoyo_tecnico TEXT,
            observaciones TEXT,
            fecha_inicio DATE NOT NULL,
            fecha_fin DATE NOT NULL,
            hora_inicio TIME,
            hora_fin TIME,
            estado TEXT DEFAULT 'Pendiente',  -- ← AGREGAR ESTA LÍNEA
            archivos_adjuntos TEXT,
            fecha_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (perito_id) REFERENCES peritos(id)
        )
    ''')
    print("✅ Tabla 'actividades_peritos' verificada/creada")





    # ============================================
    # CREAR USUARIO ADMIN POR DEFECTO
    # ============================================
    cursor.execute('SELECT COUNT(*) FROM usuarios WHERE usuario = "admin"')
    if cursor.fetchone()[0] == 0:
        # Encriptar contraseña
        password_hash = bcrypt.hashpw('admin123'.encode('utf-8'), bcrypt.gensalt())
        cursor.execute('''
            INSERT INTO usuarios (usuario, password, rol, nombre_completo, primer_inicio)
            VALUES (?, ?, ?, ?, ?)
        ''', ('admin', password_hash.decode('utf-8'), 'admin', 'Administrador del Sistema', 0))
        print("✅ Usuario 'admin' creado con contraseña 'admin123'")
    
    conn.commit()


    # Agregar columna para archivos adjuntos si no existe
    try:
        cursor.execute('ALTER TABLE asignaciones ADD COLUMN archivos_adjuntos TEXT')
        conn.commit()
        print("✅ Columna 'archivos_adjuntos' agregada a la tabla asignaciones")
    except sqlite3.OperationalError:
        print("ℹ️ Columna 'archivos_adjuntos' ya existe")
        
    # Agregar columna para hoja de envío de designación si no existe
    try:
        cursor.execute('ALTER TABLE asignaciones ADD COLUMN hoja_envio_designacion TEXT')
        conn.commit()
        print("✅ Columna 'hoja_envio_designacion' agregada a la tabla asignaciones")
    except sqlite3.OperationalError:
        print("ℹ️ Columna 'hoja_envio_designacion' ya existe")

    # Agregar columna para tipo de actividad si no existe
    try:
        cursor.execute('ALTER TABLE asignaciones ADD COLUMN tipo_actividad TEXT')
        conn.commit()
        print("✅ Columna 'tipo_actividad' agregada a la tabla asignaciones")
    except sqlite3.OperationalError:
        print("ℹ️ Columna 'tipo_actividad' ya existe")

    # Agregar columna para apoyo técnico si no existe
    try:
        cursor.execute('ALTER TABLE asignaciones ADD COLUMN apoyo_tecnico TEXT')
        conn.commit()
        print("✅ Columna 'apoyo_tecnico' agregada a la tabla asignaciones")
    except sqlite3.OperationalError:
        print("ℹ️ Columna 'apoyo_tecnico' ya existe")
    
    # Renombrar columna apoyo_tecnico a apoyos_tecnicos (para JSON array)
    try:
        cursor.execute('ALTER TABLE asignaciones RENAME COLUMN apoyo_tecnico TO apoyos_tecnicos')
        conn.commit()
        print("✅ Columna renombrada de 'apoyo_tecnico' a 'apoyos_tecnicos'")
    except sqlite3.OperationalError:
        print("ℹ️ Columna 'apoyos_tecnicos' ya existe o no se pudo renombrar")

    # Agregar columna usuario_id a archivos_asignaciones si no existe
    try:
        cursor.execute('ALTER TABLE archivos_asignaciones ADD COLUMN usuario_id INTEGER')
        conn.commit()
        print("✅ Columna 'usuario_id' agregada a archivos_asignaciones")
    except sqlite3.OperationalError:
        pass
    
    try:
        cursor.execute('ALTER TABLE archivos_asignaciones ADD COLUMN usuario_nombre TEXT')
        conn.commit()
        print("✅ Columna 'usuario_nombre' agregada a archivos_asignaciones")
    except sqlite3.OperationalError:
        pass


    conn.close()

    



# ============================================
# FUNCIONES DE AUTENTICACIÓN
# ============================================


def registrar_auditoria(usuario_id, usuario_nombre, accion, modulo, descripcion='', registro_id=None):
    """
    Registra una acción en la tabla de auditoría con IP, User-Agent y hora de Perú
    """
    try:
        conn = sqlite3.connect('database.db')
        cursor = conn.cursor()
        
        # Obtener IP y User-Agent
        ip_address = request.remote_addr if request else 'Sistema'
        user_agent = request.headers.get('User-Agent', 'Desconocido')[:200] if request else 'Sistema'
        
        # Obtener hora actual de Perú (GMT-5)
        from datetime import datetime, timedelta
        fecha_peru = (datetime.utcnow() - timedelta(hours=5)).strftime('%Y-%m-%d %H:%M:%S')
        
        # USAR NOMBRES CORRECTOS: detalles y fecha
        cursor.execute('''
            INSERT INTO auditoria (
                usuario_id, usuario_nombre, accion, modulo, 
                detalles, registro_id, ip_address, user_agent, fecha
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            usuario_id,
            usuario_nombre,
            accion,
            modulo,
            descripcion,  # Se mapea a 'detalles'
            registro_id,
            ip_address,
            user_agent,
            fecha_peru  # Se mapea a 'fecha'
        ))
        
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"❌ Error al registrar auditoría: {e}")
        import traceback
        traceback.print_exc()
        return False


def login_required(f):
    """
    Decorador para requerir login en las rutas
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    """
    Decorador para requerir rol admin
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario_id' not in session:
            return redirect(url_for('login'))
        if session.get('rol') != 'admin':
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function


def invitado_allowed(f):
    """
    Decorador para rutas accesibles por admin, perito e invitado
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario_id' not in session:
            return redirect(url_for('login'))
        # Permitir admin, perito e invitado
        if session.get('rol') not in ['admin', 'perito', 'invitado']:
            flash('No tienes permisos para acceder a esta página', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def get_usuario_actual():
    """
    Obtiene los datos del usuario actual desde la sesión
    """
    if 'usuario_id' in session:
        return {
            'id': session.get('usuario_id'),
            'usuario': session.get('usuario'),
            'rol': session.get('rol'),
            'nombre': session.get('nombre_completo'),
            'perito_id': session.get('perito_id')
        }
    return None


def actualizar_estados_automaticos():
    """
    Actualiza automáticamente los estados de las asignaciones según las fechas
    """
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    
    hoy = datetime.now().date()
    
    # Cambiar a "En Proceso" si la fecha de inicio es hoy o pasó
    cursor.execute('''
        UPDATE asignaciones 
        SET estado = "En Proceso"
        WHERE estado = "Pendiente" 
        AND date(fecha_inicio) <= date('now')
        AND date(fecha_fin) >= date('now')
    ''')
    
    # Cambiar a "Completado" si la fecha de fin ya pasó
    cursor.execute('''
        UPDATE asignaciones 
        SET estado = "Completado"
        WHERE estado IN ("Pendiente", "En Proceso")
        AND date(fecha_fin) < date('now')
    ''')
    
    conn.commit()
    conn.close()


# ============================================================================
# FUNCIONES AUXILIARES
# ============================================================================


def verificar_disponibilidad(perito_id, fecha_inicio, fecha_fin, asignacion_id=None):
    """
    Verifica si un perito está disponible en un rango de fechas.
    
    Args:
        perito_id: ID del perito a verificar
        fecha_inicio: Fecha de inicio de la asignación (formato: YYYY-MM-DD)
        fecha_fin: Fecha de fin de la asignación (formato: YYYY-MM-DD)
        asignacion_id: ID de asignación a excluir (para ediciones)
    
    Returns:
        tuple: (disponible: bool, conflictos: list de diccionarios)
    """
    conn = sqlite3.connect('database.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    conflictos = []
    
    # ==================== VERIFICAR ASIGNACIONES ====================
    query = '''
        SELECT id, hoja_envio, expediente, dependencia, tipo_perito,
               carpeta_fiscal, observaciones, lugar, fecha_inicio, fecha_fin,
               perito_asignado, estado
        FROM asignaciones
        WHERE perito_id = ?
        AND estado != 'Cancelado'
        AND (
            (fecha_inicio <= ? AND fecha_fin >= ?) OR
            (fecha_inicio <= ? AND fecha_fin >= ?) OR
            (fecha_inicio >= ? AND fecha_fin <= ?)
        )
    '''
    
    params = [
        perito_id,
        fecha_fin, fecha_inicio,
        fecha_fin, fecha_fin,
        fecha_inicio, fecha_fin
    ]
    
    # Excluir la asignación actual si estamos editando
    if asignacion_id:
        query += ' AND id != ?'
        params.append(asignacion_id)
    
    cursor.execute(query, params)
    resultados_asignaciones = cursor.fetchall()
    
    # Convertir asignaciones a diccionarios
    for row in resultados_asignaciones:
        conflictos.append({
            'id': row['id'],
            'tipo': 'asignacion',
            'hoja_envio': row['hoja_envio'],
            'expediente': row['expediente'],
            'dependencia': row['dependencia'],
            'tipo_perito': row['tipo_perito'],
            'carpeta_fiscal': row['carpeta_fiscal'],
            'observaciones': row['observaciones'],
            'lugar': row['lugar'],
            'fecha_inicio': row['fecha_inicio'],
            'fecha_fin': row['fecha_fin'],
            'perito_asignado': row['perito_asignado'],
            'estado': row['estado']
        })
    
    # ==================== VERIFICAR VACACIONES ====================
    cursor.execute('''
        SELECT id, tipo_perito, fecha_inicio, fecha_fin, tipo_vacaciones, dias_totales, observaciones
        FROM vacaciones
        WHERE perito_id = ?
        AND (
            (fecha_inicio <= ? AND fecha_fin >= ?) OR
            (fecha_inicio <= ? AND fecha_fin >= ?) OR
            (fecha_inicio >= ? AND fecha_fin <= ?)
        )
    ''', (perito_id, fecha_fin, fecha_inicio, fecha_fin, fecha_fin, fecha_inicio, fecha_fin))
    
    resultados_vacaciones = cursor.fetchall()
    
    # Convertir vacaciones a diccionarios
    for row in resultados_vacaciones:
        conflictos.append({
            'id': row['id'],
            'tipo': 'vacacion',
            'expediente': '🏖️ VACACIONES',
            'fecha_inicio': row['fecha_inicio'],
            'fecha_fin': row['fecha_fin'],
            'tipo_vacaciones': row['tipo_vacaciones'],
            'dias_totales': row['dias_totales'],
            'observaciones': f"Período de vacaciones ({row['tipo_vacaciones']}) - {row['dias_totales']} días"
        })
    
    conn.close()  # ← CERRAR CONEXIÓN AL FINAL
    
    disponible = len(conflictos) == 0
    
    return disponible, conflictos

def registrar_historial(asignacion_id, accion, detalles=''):
    """
    Registra una acción en el historial para auditoría.
    
    Args:
        asignacion_id: ID de la asignación relacionada
        accion: Tipo de acción (Creado, Modificado, Completado, etc.)
        detalles: Información adicional sobre la acción
    """
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    
    cursor.execute(
        'INSERT INTO historial (asignacion_id, accion, detalles) VALUES (?, ?, ?)',
        (asignacion_id, accion, detalles)
    )
    
    conn.commit()
    conn.close()


# ============================================
# RUTAS DE AUTENTICACIÓN
# ============================================

@app.route('/login', methods=['GET', 'POST'])
def login():
    """
    Página de inicio de sesión
    """
    # Si ya está logueado, redirigir al dashboard
    if 'usuario_id' in session:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        data = request.json if request.is_json else request.form
        usuario = data.get('usuario', '').strip()
        password = data.get('password', '').strip()
        
        if not usuario or not password:
            if request.is_json:
                return jsonify({'success': False, 'error': 'Usuario y contraseña son requeridos'}), 400
            return render_template('auth/login.html', error='Usuario y contraseña son requeridos')
        
        conn = sqlite3.connect('database.db')
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute('SELECT * FROM usuarios WHERE usuario = ? AND activo = 1', (usuario,))
        user = cursor.fetchone()
        
        if user and bcrypt.checkpw(password.encode('utf-8'), user['password'].encode('utf-8')):
            # Login exitoso
            session.permanent = True
            session['usuario_id'] = user['id']
            session['usuario'] = user['usuario']
            session['rol'] = user['rol']
            session['nombre_completo'] = user['nombre_completo']
            session['perito_id'] = user['perito_id']
            session['primer_inicio'] = user['primer_inicio']
            
            # Actualizar último acceso
            cursor.execute('UPDATE usuarios SET ultimo_acceso = CURRENT_TIMESTAMP WHERE id = ?', (user['id'],))
            conn.commit()
            
            # Registrar en auditoría
            registrar_auditoria(
                user['id'], 
                user['nombre_completo'], 
                'LOGIN', 
                'AUTENTICACIÓN', 
                f"Inicio de sesión exitoso desde IP: {request.remote_addr}"
            )
            
            conn.close()
            
            # Si es primer inicio, redirigir a cambio de contraseña
            if user['primer_inicio'] == 1:
                if request.is_json:
                    return jsonify({'success': True, 'redirect': '/cambiar-password', 'primer_inicio': True})
                return redirect(url_for('cambiar_password'))
            
            if request.is_json:
                return jsonify({'success': True, 'redirect': '/'})
            return redirect(url_for('index'))
        else:
            conn.close()
            # Registrar intento fallido (sin usuario válido)
            if request.is_json:
                return jsonify({'success': False, 'error': 'Usuario o contraseña incorrectos'}), 401
            return render_template('auth/login.html', error='Usuario o contraseña incorrectos')
    
    return render_template('auth/login.html')

@app.route('/logout')
def logout():
    """
    Cerrar sesión
    """
    if 'usuario_id' in session:
        registrar_auditoria(
            session['usuario_id'],
            session.get('nombre_completo', 'Usuario'),
            'LOGOUT',
            'AUTENTICACIÓN',
            'Cierre de sesión'
        )
    
    session.clear()
    return redirect(url_for('login'))

@app.route('/cambiar-password', methods=['GET', 'POST'])
@login_required
def cambiar_password():
    """
    Página para cambiar contraseña
    """
    if request.method == 'POST':
        data = request.json if request.is_json else request.form
        password_actual = data.get('password_actual', '').strip()
        password_nueva = data.get('password_nueva', '').strip()
        password_confirmar = data.get('password_confirmar', '').strip()
        
        # Validaciones
        if not password_nueva or not password_confirmar:
            error = 'Todos los campos son requeridos'
            if request.is_json:
                return jsonify({'success': False, 'error': error}), 400
            return render_template('cambiar_password.html', error=error)
        
        if password_nueva != password_confirmar:
            error = 'Las contraseñas no coinciden'
            if request.is_json:
                return jsonify({'success': False, 'error': error}), 400
            return render_template('cambiar_password.html', error=error)
        
        if len(password_nueva) < 6:
            error = 'La contraseña debe tener al menos 6 caracteres'
            if request.is_json:
                return jsonify({'success': False, 'error': error}), 400
            return render_template('cambiar_password.html', error=error)
        
        conn = sqlite3.connect('database.db')
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute('SELECT password FROM usuarios WHERE id = ?', (session['usuario_id'],))
        user = cursor.fetchone()
        
        # Si NO es primer inicio, validar contraseña actual
        if session.get('primer_inicio') != 1:
            if not password_actual:
                conn.close()
                error = 'Debe ingresar su contraseña actual'
                if request.is_json:
                    return jsonify({'success': False, 'error': error}), 400
                return render_template('cambiar_password.html', error=error)
            
            if not bcrypt.checkpw(password_actual.encode('utf-8'), user['password'].encode('utf-8')):
                conn.close()
                error = 'La contraseña actual es incorrecta'
                if request.is_json:
                    return jsonify({'success': False, 'error': error}), 400
                return render_template('cambiar_password.html', error=error)
        
        # Actualizar contraseña
        password_hash = bcrypt.hashpw(password_nueva.encode('utf-8'), bcrypt.gensalt())
        cursor.execute('''
            UPDATE usuarios SET password = ?, primer_inicio = 0 WHERE id = ?
        ''', (password_hash.decode('utf-8'), session['usuario_id']))
        
        conn.commit()
        conn.close()
        
        # Actualizar sesión
        session['primer_inicio'] = 0
        
        # Registrar en auditoría
        registrar_auditoria(
            session['usuario_id'],
            session.get('nombre_completo', 'Usuario'),
            'CAMBIO_PASSWORD',
            'AUTENTICACIÓN',
            'Cambio de contraseña exitoso'
        )
        
        if request.is_json:
            return jsonify({'success': True, 'message': 'Contraseña actualizada correctamente', 'redirect': '/'})
        return redirect(url_for('index'))
    
    return render_template('auth/cambiar_password.html', primer_inicio=session.get('primer_inicio', 0))   



# ============================================
# GESTIÓN DE USUARIOS (Solo Admin)
# ============================================

@app.route('/usuarios')
@admin_required
def gestion_usuarios():
    """
    Página de gestión de usuarios
    """
    conn = sqlite3.connect('database.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # Obtener todos los usuarios
    cursor.execute('''
        SELECT u.*, p.nombre_completo as perito_nombre
        FROM usuarios u
        LEFT JOIN peritos p ON u.perito_id = p.id
        ORDER BY u.fecha_creacion DESC
    ''')
    
    usuarios = []
    for row in cursor.fetchall():
        usuarios.append({
            'id': row['id'],
            'usuario': row['usuario'],
            'rol': row['rol'],
            'perito_id': row['perito_id'],
            'perito_nombre': row['perito_nombre'],
            'nombre_completo': row['nombre_completo'],
            'email': row['email'],
            'activo': row['activo'],
            'primer_inicio': row['primer_inicio'],
            'fecha_creacion': row['fecha_creacion'],
            'ultimo_acceso': row['ultimo_acceso']
        })
    
    # Obtener peritos para el select
    
    
    cursor.execute('SELECT id, nombre_completo, tipo FROM peritos WHERE estado = "Activo" ORDER BY tipo, nombre_completo')
    peritos = [{'id': row['id'], 'nombre': row['nombre_completo'], 'tipo': row['tipo']} for row in cursor.fetchall()]

    conn.close()
    
    return render_template('admin/usuarios.html', usuarios=usuarios, peritos=peritos)

@app.route('/api/usuario', methods=['POST'])
@admin_required
def crear_usuario():
    """
    Crear nuevo usuario
    """
    data = request.json
    
    # Validaciones
    if not data.get('usuario') or not data.get('password') or not data.get('rol'):
        return jsonify({'success': False, 'error': 'Usuario, contraseña y rol son requeridos'}), 400
    
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    
    # Verificar si el usuario ya existe
    cursor.execute('SELECT id FROM usuarios WHERE usuario = ?', (data['usuario'],))
    if cursor.fetchone():
        conn.close()
        return jsonify({'success': False, 'error': 'El nombre de usuario ya existe'}), 409
    
    # Encriptar contraseña
    password_hash = bcrypt.hashpw(data['password'].encode('utf-8'), bcrypt.gensalt())
    
    # Insertar usuario
    cursor.execute('''
        INSERT INTO usuarios (usuario, password, rol, perito_id, nombre_completo, email, activo, primer_inicio)
        VALUES (?, ?, ?, ?, ?, ?, 1, 1)
    ''', (
        data['usuario'],
        password_hash.decode('utf-8'),
        data['rol'],
        data.get('perito_id') if data.get('perito_id') else None,
        data.get('nombre_completo', ''),
        data.get('email', '')
    ))
    
    usuario_id = cursor.lastrowid
    conn.commit()
    conn.close()
    
    # Registrar en auditoría
    registrar_auditoria(
        session['usuario_id'],
        session.get('nombre_completo'),
        'CREAR_USUARIO',
        'USUARIOS',
        f"Usuario '{data['usuario']}' creado con rol '{data['rol']}'",
        usuario_id
    )
    
    return jsonify({'success': True, 'id': usuario_id, 'message': 'Usuario creado exitosamente'})

@app.route('/api/usuario/<int:id>', methods=['PUT'])
@admin_required
def actualizar_usuario(id):
    """
    Actualizar usuario existente
    """
    data = request.json
    
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    
    # Verificar que el usuario existe
    cursor.execute('SELECT id FROM usuarios WHERE id = ?', (id,))
    if not cursor.fetchone():
        conn.close()
        return jsonify({'success': False, 'error': 'Usuario no encontrado'}), 404
    
    # Construir query de actualización
    campos = []
    valores = []
    
    if 'nombre_completo' in data:
        campos.append('nombre_completo = ?')
        valores.append(data['nombre_completo'])
    
    if 'email' in data:
        campos.append('email = ?')
        valores.append(data['email'])
    
    if 'rol' in data:
        campos.append('rol = ?')
        valores.append(data['rol'])
    
    if 'perito_id' in data:
        campos.append('perito_id = ?')
        valores.append(data['perito_id'] if data['perito_id'] else None)
    
    if 'activo' in data:
        campos.append('activo = ?')
        valores.append(1 if data['activo'] else 0)
    
    if not campos:
        conn.close()
        return jsonify({'success': False, 'error': 'No hay campos para actualizar'}), 400
    
    valores.append(id)
    query = f"UPDATE usuarios SET {', '.join(campos)} WHERE id = ?"
    
    cursor.execute(query, valores)
    conn.commit()
    conn.close()
    
    # Registrar en auditoría
    registrar_auditoria(
        session['usuario_id'],
        session.get('nombre_completo'),
        'EDITAR_USUARIO',
        'USUARIOS',
        f"Usuario ID {id} actualizado",
        id
    )
    
    return jsonify({'success': True, 'message': 'Usuario actualizado exitosamente'})

@app.route('/api/usuario/<int:id>/resetear-password', methods=['POST'])
@admin_required
def resetear_password(id):
    """
    Resetear contraseña de un usuario
    """
    data = request.json
    nueva_password = data.get('password', 'password123')
    
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    
    # Verificar que el usuario existe
    cursor.execute('SELECT usuario FROM usuarios WHERE id = ?', (id,))
    user = cursor.fetchone()
    if not user:
        conn.close()
        return jsonify({'success': False, 'error': 'Usuario no encontrado'}), 404
    
    # Encriptar nueva contraseña
    password_hash = bcrypt.hashpw(nueva_password.encode('utf-8'), bcrypt.gensalt())
    
    # Actualizar contraseña y marcar primer_inicio
    cursor.execute('''
        UPDATE usuarios SET password = ?, primer_inicio = 1 WHERE id = ?
    ''', (password_hash.decode('utf-8'), id))
    
    conn.commit()
    conn.close()
    
    # Registrar en auditoría
    registrar_auditoria(
        session['usuario_id'],
        session.get('nombre_completo'),
        'RESETEAR_PASSWORD',
        'USUARIOS',
        f"Contraseña reseteada para usuario '{user[0]}'",
        id
    )
    
    return jsonify({'success': True, 'message': f'Contraseña reseteada. Nueva contraseña: {nueva_password}'})

@app.route('/api/usuario/<int:id>', methods=['DELETE'])
@admin_required
def eliminar_usuario(id):
    """
    Eliminar usuario
    """
    # No permitir eliminar el propio usuario
    if id == session.get('usuario_id'):
        return jsonify({'success': False, 'error': 'No puedes eliminar tu propio usuario'}), 400
    
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    
    # Verificar que el usuario existe
    cursor.execute('SELECT usuario FROM usuarios WHERE id = ?', (id,))
    user = cursor.fetchone()
    if not user:
        conn.close()
        return jsonify({'success': False, 'error': 'Usuario no encontrado'}), 404
    
    # Eliminar usuario
    cursor.execute('DELETE FROM usuarios WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    
    # Registrar en auditoría
    registrar_auditoria(
        session['usuario_id'],
        session.get('nombre_completo'),
        'ELIMINAR_USUARIO',
        'USUARIOS',
        f"Usuario '{user[0]}' eliminado",
        id
    )
    
    return jsonify({'success': True, 'message': 'Usuario eliminado exitosamente'})



@app.route('/auditoria')
@admin_required
def auditoria():
    """
    Página de auditoría - solo para administradores
    """
    conn = sqlite3.connect('database.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # Obtener filtros
    filtro_usuario = request.args.get('usuario', '')
    filtro_accion = request.args.get('accion', '')
    filtro_modulo = request.args.get('modulo', '')
    filtro_fecha_desde = request.args.get('fecha_desde', '')
    filtro_fecha_hasta = request.args.get('fecha_hasta', '')
    
    # PAGINACIÓN
    page = int(request.args.get('page', 1))
    per_page = 50  # Registros por página
    offset = (page - 1) * per_page
    
    # Construir query base
    query_count = 'SELECT COUNT(*) as total FROM auditoria WHERE 1=1'
    query = 'SELECT * FROM auditoria WHERE 1=1'
    params = []
    
    if filtro_usuario:
        query += ' AND usuario_nombre LIKE ?'
        query_count += ' AND usuario_nombre LIKE ?'
        params.append(f'%{filtro_usuario}%')
    
    if filtro_accion:
        query += ' AND accion = ?'
        query_count += ' AND accion = ?'
        params.append(filtro_accion)
    
    if filtro_modulo:
        query += ' AND modulo = ?'
        query_count += ' AND modulo = ?'
        params.append(filtro_modulo)
    
    if filtro_fecha_desde:
        query += ' AND DATE(fecha) >= ?'  # USAR 'fecha'
        query_count += ' AND DATE(fecha) >= ?'
        params.append(filtro_fecha_desde)
    
    if filtro_fecha_hasta:
        query += ' AND DATE(fecha) <= ?'  # USAR 'fecha'
        query_count += ' AND DATE(fecha) <= ?'
        params.append(filtro_fecha_hasta)
    
    try:
        # Obtener total de registros
        cursor.execute(query_count, params)
        total_registros = cursor.fetchone()['total']
        total_paginas = (total_registros + per_page - 1) // per_page
        
        # Obtener registros de la página actual
        query += f' ORDER BY fecha DESC LIMIT {per_page} OFFSET {offset}'  # USAR 'fecha'
        cursor.execute(query, params)
        
        registros = []
        for row in cursor.fetchall():
            registros.append({
                'id': row['id'],
                'fecha': row['fecha'],  # USAR 'fecha'
                'usuario_nombre': row['usuario_nombre'],
                'accion': row['accion'],
                'modulo': row['modulo'],
                'detalles': row['detalles'] if row['detalles'] else '',  # USAR 'detalles'
                'ip_address': row['ip_address'] if row['ip_address'] else 'N/A',
                'user_agent': row['user_agent'] if row['user_agent'] else 'N/A'
            })
        


        # DEBUG: Ver cuántos registros se enviaron
        print(f"🔍 DEBUG: Se encontraron {len(registros)} registros")
        if registros:
            print(f"🔍 DEBUG: Primer registro: {registros[0]}")

        # Obtener acciones únicas
        cursor.execute('SELECT DISTINCT accion FROM auditoria ORDER BY accion')
        acciones = [row['accion'] for row in cursor.fetchall()]
        
        # Obtener módulos únicos
        cursor.execute('SELECT DISTINCT modulo FROM auditoria ORDER BY modulo')
        modulos = [row['modulo'] for row in cursor.fetchall()]
        
        conn.close()
        
        return render_template('admin/auditoria.html', 
                              registros=registros,
                              acciones=acciones,
                              modulos=modulos,
                              filtros={
                                  'usuario': filtro_usuario,
                                  'accion': filtro_accion,
                                  'modulo': filtro_modulo,
                                  'fecha_desde': filtro_fecha_desde,
                                  'fecha_hasta': filtro_fecha_hasta
                              },
                              page=page,
                              total_paginas=total_paginas,
                              total_registros=total_registros)
    except Exception as e:
        conn.close()
        print(f"❌ Error en auditoría: {e}")
        import traceback
        traceback.print_exc()
        return f"Error al cargar auditoría: {str(e)}", 500



@app.route('/api/auditoria/exportar')
@admin_required
def exportar_auditoria():
    """
    Exporta la auditoría a Excel con formato
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    
    conn = sqlite3.connect('database.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # Obtener todos los registros - USAR 'fecha'
    cursor.execute('SELECT * FROM auditoria ORDER BY fecha DESC')
    
    # Crear libro Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Auditoría"
    
    # Estilo de cabeceras
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Cabeceras
    headers = ['ID', 'Fecha/Hora', 'Usuario', 'Acción', 'Módulo', 'Detalles', 'IP', 'Navegador']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos - USAR 'fecha' y 'detalles'
    for row_idx, row in enumerate(cursor.fetchall(), start=2):
        ws.cell(row=row_idx, column=1, value=row['id'])
        ws.cell(row=row_idx, column=2, value=row['fecha'])  # CAMBIO AQUÍ
        ws.cell(row=row_idx, column=3, value=row['usuario_nombre'])
        ws.cell(row=row_idx, column=4, value=row['accion'])
        ws.cell(row=row_idx, column=5, value=row['modulo'])
        ws.cell(row=row_idx, column=6, value=row['detalles'] or '')  # CAMBIO AQUÍ
        ws.cell(row=row_idx, column=7, value=row['ip_address'] if row['ip_address'] else 'N/A')
        ws.cell(row=row_idx, column=8, value=row['user_agent'][:50] if row['user_agent'] else 'N/A')
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 50
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 40
    
    conn.close()
    
    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Crear respuesta
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'auditoria_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

# ============================================================================
# RUTAS PRINCIPALES
# ============================================================================
#------------------------------------------------------------------------------------------------------- comienza aagregar nuevos codigos par apginacion
# NUEVO CODIGO CAMBIADO A ROLES

@app.route('/')
@invitado_allowed
def index():
    """
    Página principal - Dashboard con estadísticas generales y paginación
    Filtra según el rol del usuario
    """
    # Actualizar estados automáticamente
    actualizar_estados_automaticos()
    
    # Obtener número de página actual (por defecto página 1)
    pagina = request.args.get('pagina', 1, type=int)
    por_pagina = 10
    
    conn = sqlite3.connect('database.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # Determinar si filtrar por perito
    perito_id = session.get('perito_id')
    es_admin = session.get('rol') == 'admin'
    es_invitado = session.get('rol') == 'invitado'  # ← AGREGAR
    
    # Obtener info del perito si es usuario perito
    tipo_perito = None
    if perito_id and not es_admin and not es_invitado:  # ← MODIFICAR
        cursor.execute('SELECT tipo FROM peritos WHERE id = ?', (perito_id,))
        result = cursor.fetchone()
        if result:
            tipo_perito = result['tipo']
    
    # ==================== ESTADÍSTICAS ====================
    if es_admin or es_invitado:  # ← ADMIN E INVITADO VEN TODO
        cursor.execute('SELECT COUNT(*) FROM asignaciones WHERE estado != "Cancelado"')
        total_asignaciones = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(*) FROM asignaciones WHERE estado = "Pendiente"')
        pendientes = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(*) FROM asignaciones WHERE estado = "En Proceso"')
        en_proceso = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(*) FROM asignaciones WHERE estado = "Completado"')
        completados = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(*) FROM asignaciones WHERE estado = "Cancelado"')
        cancelados = cursor.fetchone()[0]
    else:  # PERITO VE SOLO LO SUYO
        cursor.execute('SELECT COUNT(*) FROM asignaciones WHERE estado != "Cancelado" AND perito_id = ?', (perito_id,))
        total_asignaciones = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(*) FROM asignaciones WHERE estado = "Pendiente" AND perito_id = ?', (perito_id,))
        pendientes = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(*) FROM asignaciones WHERE estado = "En Proceso" AND perito_id = ?', (perito_id,))
        en_proceso = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(*) FROM asignaciones WHERE estado = "Completado" AND perito_id = ?', (perito_id,))
        completados = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(*) FROM asignaciones WHERE estado = "Cancelado" AND perito_id = ?', (perito_id,))
        cancelados = cursor.fetchone()[0]
    
    # ==================== ASIGNACIONES ====================
    offset = (pagina - 1) * por_pagina
    
    if es_admin or es_invitado:  # ← ADMIN E INVITADO VEN TODO
        cursor.execute('''
            SELECT 
                a.id, a.hoja_envio, a.expediente, a.dependencia, 
                a.tipo_perito, a.tipo_actividad, a.carpeta_fiscal, a.observaciones, a.lugar,
                a.fecha_inicio, a.fecha_fin, a.perito_asignado, a.perito_id,
                a.desginacion, a.oficio_desplazamiento, a.estado, a.fecha_registro,
                p.nombre_completo
            FROM asignaciones a
            LEFT JOIN peritos p ON a.perito_id = p.id
            WHERE a.estado != "Cancelado"
            ORDER BY a.fecha_registro DESC
            LIMIT ? OFFSET ?
        ''', (por_pagina, offset))
    else:  # PERITO VE SOLO LO SUYO
        cursor.execute('''
            SELECT 
                a.id, a.hoja_envio, a.expediente, a.dependencia, 
                a.tipo_perito, a.tipo_actividad, a.carpeta_fiscal, a.observaciones, a.lugar,
                a.fecha_inicio, a.fecha_fin, a.perito_asignado, a.perito_id,
                a.desginacion, a.oficio_desplazamiento, a.estado, a.fecha_registro,
                p.nombre_completo
            FROM asignaciones a
            LEFT JOIN peritos p ON a.perito_id = p.id
            WHERE a.estado != "Cancelado" AND a.perito_id = ?
            ORDER BY a.fecha_registro DESC
            LIMIT ? OFFSET ?
        ''', (perito_id, por_pagina, offset))
    
    asignaciones_recientes = []
    for row in cursor.fetchall():
        asignaciones_recientes.append({
            'id': row['id'],
            'expediente': row['expediente'],
            'carpeta_fiscal': row['carpeta_fiscal'],
            'dependencia': row['dependencia'],
            'tipo_perito': row['tipo_perito'],
            'tipo_actividad': row['tipo_actividad'],
            'fecha_inicio': row['fecha_inicio'],
            'fecha_fin': row['fecha_fin'],
            'perito': row['nombre_completo'] if row['nombre_completo'] else row['perito_asignado'],
            'estado': row['estado'],
            'lugar': row['lugar']
        })
    
    # Calcular total de páginas
    total_paginas = (total_asignaciones + por_pagina - 1) // por_pagina if total_asignaciones > 0 else 1
    
    conn.close()
    
    return render_template('core/index.html',
                         total=total_asignaciones,
                         pendientes=pendientes,
                         en_proceso=en_proceso,
                         completados=completados,
                         cancelados=cancelados,
                         asignaciones=asignaciones_recientes,
                         pagina_actual=pagina,
                         total_paginas=total_paginas,
                         por_pagina=por_pagina,
                         es_admin=es_admin,
                         tipo_perito=tipo_perito)
#CAMBIOS ---
    

@app.route('/nuevo')
@login_required  # ← AGREGAR
def nuevo():
    """
    Página para registrar nueva asignación
    """
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    
    # Obtener lista de peritos activos
    cursor.execute('SELECT id, nombre_completo, tipo FROM peritos WHERE estado = "Activo" ORDER BY tipo, nombre_completo')
    peritos = cursor.fetchall()
    conn.close()
    
    # Organizar peritos por tipo
    peritos_por_tipo = {
        'Informático': [],
        'Acústico': [],
        'Antropólogo': [],
        'Contable': []
    }
    
    for perito in peritos:
        peritos_por_tipo[perito[2]].append({
            'id': perito[0],
            'nombre': perito[1]
        })
    
    return render_template('core/nuevo.html', peritos=peritos_por_tipo)

@app.route('/buscar')
@invitado_allowed  # ← CAMBIAR
def buscar():
    """
    Página de búsqueda de asignaciones
    """
    tipo_perito = None
    nombre_perito = None
    perito_id = session.get('perito_id')
    es_admin = session.get('rol') == 'admin'
    
    if perito_id and not es_admin:
        conn = sqlite3.connect('database.db')
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute('SELECT tipo, nombre_completo FROM peritos WHERE id = ?', (perito_id,))
        result = cursor.fetchone()
        if result:
            tipo_perito = result['tipo']
            nombre_perito = result['nombre_completo']
        conn.close()
    
    return render_template('core/buscar.html', 
                          es_admin=es_admin, 
                          tipo_perito=tipo_perito,
                          nombre_perito=nombre_perito,
                          perito_id=perito_id)

#@app.route('/calendario')
#@login_required  # ← AGREGAR
#def calendario():
#    """
#    Vista de calendario con asignaciones
#    """
#    return render_template('core/calendario.html')

@app.route('/calendario')
@invitado_allowed  # ← CAMBIAR
def calendario():
    """
    Vista de calendario con asignaciones
    """
    tipo_perito = None
    nombre_perito = None
    perito_id = session.get('perito_id')
    es_admin = session.get('rol') == 'admin'
    
    if perito_id and not es_admin:
        conn = sqlite3.connect('database.db')
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute('SELECT tipo, nombre_completo FROM peritos WHERE id = ?', (perito_id,))
        result = cursor.fetchone()
        if result:
            tipo_perito = result['tipo']
            nombre_perito = result['nombre_completo']
        conn.close()
    
    return render_template('core/calendario.html',
                          es_admin=es_admin,
                          tipo_perito=tipo_perito,
                          nombre_perito=nombre_perito,
                          perito_id=perito_id)

@app.route('/peritos')
@login_required  # ← AGREGAR
def peritos():
    """
    Gestión de peritos
    """
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    
    cursor.execute('SELECT * FROM peritos ORDER BY tipo, nombre_completo')
    peritos_list = []
    
    for row in cursor.fetchall():
        # Contar asignaciones del perito
        cursor.execute('SELECT COUNT(*) FROM asignaciones WHERE perito_id = ?', (row[0],))
        total_asignaciones = cursor.fetchone()[0]
        
        peritos_list.append({
            'id': row[0],
            'nombre': row[1],
            'tipo': row[2],
            'estado': row[3],
            'total_asignaciones': total_asignaciones
        })
    
    conn.close()
    
    return render_template('core/peritos.html', peritos=peritos_list)

@app.route('/reportes')
@login_required  # ← AGREGAR
def reportes():
    """
    Página de reportes y estadísticas
    """
    return render_template('core/reportes.html')

# ============================================================================
# API ENDPOINTS
# ============================================================================

@app.route('/api/asignaciones', methods=['GET'])
@login_required
def get_asignaciones():
    """
    Obtiene asignaciones y vacaciones con filtros
    Filtra según el rol del usuario
    """
    conn = sqlite3.connect('database.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # Determinar si filtrar por perito
    perito_id = session.get('perito_id')
    es_admin = session.get('rol') == 'admin'
    
    # ==================== OBTENER ASIGNACIONES ====================
    query = '''
        SELECT a.*, p.nombre_completo
        FROM asignaciones a
        LEFT JOIN peritos p ON a.perito_id = p.id
        WHERE a.estado != 'Cancelado'
    '''
    params = []
    
    # Filtro por rol (perito solo ve las suyas)
    if not es_admin and perito_id:
        query += ' AND a.perito_id = ?'
        params.append(perito_id)
    
    # Filtro por estado
    if request.args.get('estado'):
        query += ' AND a.estado = ?'
        params.append(request.args.get('estado'))
    
    # Filtro por perito (para admin)
    if request.args.get('perito_id') and es_admin:
        query += ' AND a.perito_id = ?'
        params.append(request.args.get('perito_id'))
    
    # Filtro por rango de fechas
    if request.args.get('fecha_desde'):
        query += ' AND a.fecha_inicio >= ?'
        params.append(request.args.get('fecha_desde'))
    
    if request.args.get('fecha_hasta'):
        query += ' AND a.fecha_fin <= ?'
        params.append(request.args.get('fecha_hasta'))
    
    query += ' ORDER BY a.fecha_inicio DESC'
    
    cursor.execute(query, params)
    
    resultados = []
    
    for row in cursor.fetchall():
        resultados.append({
            'id': row['id'],
            'tipo': 'asignacion',
            'hoja_envio': row['hoja_envio'],
            'expediente': row['expediente'],
            'dependencia': row['dependencia'],
            'tipo_perito': row['tipo_perito'],
            'tipo_actividad': row['tipo_actividad'],
            'apoyos_tecnicos': row['apoyo_tecnico'],
            'carpeta_fiscal': row['carpeta_fiscal'],
            'hoja_envio_designacion': row['hoja_envio_designacion'],
            'observaciones': row['observaciones'],
            'lugar': row['lugar'],
            'fecha_inicio': row['fecha_inicio'],
            'fecha_fin': row['fecha_fin'],
            'perito_asignado': row['perito_asignado'],
            'desginacion': row['desginacion'],
            'oficio_desplazamiento': row['oficio_desplazamiento'],
            'estado': row['estado'],
            'perito_nombre': row['nombre_completo']
        })
    

    
    # ==================== OBTENER VACACIONES ====================
    query_vac = '''
        SELECT v.*, p.nombre_completo
        FROM vacaciones v
        LEFT JOIN peritos p ON v.perito_id = p.id
        WHERE 1=1
    '''
    params_vac = []
    
    
    # Filtro por rol (perito solo ve las suyas)
    if not es_admin and perito_id:
        query_vac += ' AND v.perito_id = ?'
        params_vac.append(perito_id)
    
    # Filtro por perito (para admin)
    if request.args.get('perito_id') and es_admin:
        query_vac += ' AND v.perito_id = ?'
        params_vac.append(request.args.get('perito_id'))
    
    # Filtro por rango de fechas
    if request.args.get('fecha_desde'):
        query_vac += ' AND v.fecha_inicio >= ?'
        params_vac.append(request.args.get('fecha_desde'))
    
    if request.args.get('fecha_hasta'):
        query_vac += ' AND v.fecha_fin <= ?'
        params_vac.append(request.args.get('fecha_hasta'))
    
    query_vac += ' ORDER BY v.fecha_inicio DESC'
    
    cursor.execute(query_vac, params_vac)
    
    for row in cursor.fetchall():
        resultados.append({
            'id': f"vac_{row['id']}",
            'vacacion_id': row['id'],
            'tipo': 'vacacion',
            'perito_nombre': row['nombre_completo'],
            'tipo_perito': row['tipo_perito'],
            'fecha_inicio': row['fecha_inicio'],
            'fecha_fin': row['fecha_fin'],
            'dias_totales': row['dias_totales'],
            'tipo_vacaciones': row['tipo_vacaciones'],
            'estado': row['estado'],
            'observaciones': row['observaciones']
        })
    

    # ==================== OBTENER ACTIVIDADES DE PERITOS ====================
    query_act = '''
        SELECT a.*, p.nombre_completo
        FROM actividades_peritos a
        LEFT JOIN peritos p ON a.perito_id = p.id
        WHERE 1=1
    '''
    params_act = []
    
    # Filtro por rol (perito solo ve las suyas)
    if not es_admin and perito_id:
        query_act += ' AND a.perito_id = ?'
        params_act.append(perito_id)
    
    # Filtro por perito específico (para admin)
    if request.args.get('perito_id') and es_admin:
        query_act += ' AND a.perito_id = ?'
        params_act.append(request.args.get('perito_id'))
    
    # Filtro por rango de fechas
    if request.args.get('fecha_desde'):
        query_act += ' AND a.fecha_inicio >= ?'
        params_act.append(request.args.get('fecha_desde'))
    
    if request.args.get('fecha_hasta'):
        query_act += ' AND a.fecha_fin <= ?'
        params_act.append(request.args.get('fecha_hasta'))
    
    query_act += ' ORDER BY a.fecha_inicio DESC'
    
    cursor.execute(query_act, params_act)
    
    for row in cursor.fetchall():
        resultados.append({
            'id': f"act_{row['id']}",
            'actividad_id': row['id'],
            'tipo': 'actividad',
            'perito_nombre': row['nombre_completo'],
            'tipo_perito': row['tipo_perito'],
            'carpeta_fiscal': row['carpeta_fiscal'],
            'tipo_actividad': row['tipo_actividad'],
            'fecha_inicio': row['fecha_inicio'],
            'fecha_fin': row['fecha_fin'],
            'hora_inicio': row['hora_inicio'],
            'hora_fin': row['hora_fin'],
            'observaciones': row['observaciones'],
            'dependencia': row['dependencia'],
            'denominacion': row['denominacion']
        })

    conn.close()
    return jsonify(resultados)

 

@app.route('/api/asignacion/<int:id>', methods=['GET'])
def get_asignacion(id):
    """
    Obtiene una asignación específica por ID
    """
    conn = sqlite3.connect('database.db')
    conn.row_factory = sqlite3.Row  # ← ESTO ES CLAVE
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT a.*, p.nombre_completo
        FROM asignaciones a
        LEFT JOIN peritos p ON a.perito_id = p.id
        WHERE a.id = ?
    ''', (id,))
    
    row = cursor.fetchone()
    conn.close()
    
    if row:
        asignacion = {
            'id': row['id'],
            'hoja_envio': row['hoja_envio'],
            'expediente': row['expediente'],
            'dependencia': row['dependencia'],
            'tipo_perito': row['tipo_perito'],
            'tipo_actividad': row['tipo_actividad'],  # ← NUEVO
            'apoyos_tecnicos': row['apoyo_tecnico'],  # ← NUEVO
            'carpeta_fiscal': row['carpeta_fiscal'],
            'denominacion': row['denominacion'],  # ← AGREGAR
            'hoja_envio_designacion': row['hoja_envio_designacion'],  # ← NUEVO
            'observaciones': row['observaciones'],
            'lugar': row['lugar'],
            'fecha_inicio': row['fecha_inicio'],
            'fecha_fin': row['fecha_fin'],
            'perito_asignado': row['perito_asignado'],
            'perito_id': row['perito_id'],
            'desginacion': row['desginacion'],
            'oficio_desplazamiento': row['oficio_desplazamiento'],
            'estado': row['estado'],
            'perito_nombre': row['nombre_completo'] if row['nombre_completo'] else row['perito_asignado']
        }
        return jsonify(asignacion)
    else:
        return jsonify({'error': 'Asignación no encontrada'}), 404




@app.route('/api/asignacion', methods=['POST'])
@login_required
def crear_asignacion():
    """
    Crea una nueva asignación
    """
    data = request.json
    
    # Validar datos requeridos
    if not all(k in data for k in ('perito_id', 'fecha_inicio', 'fecha_fin')):
        return jsonify({'error': 'Faltan datos requeridos'}), 400
    
    # Verificar disponibilidad del perito
    disponible, conflictos = verificar_disponibilidad(
        data['perito_id'],
        data['fecha_inicio'],
        data['fecha_fin']
    )
    
    if not disponible:
        return jsonify({
            'error': 'El perito no está disponible en estas fechas',
            'conflictos': conflictos
        }), 409
    
    # Insertar asignación
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    
    cursor.execute('''
        INSERT INTO asignaciones (
            hoja_envio, expediente, dependencia, tipo_perito, tipo_actividad, apoyo_tecnico,
            carpeta_fiscal, denominacion, hoja_envio_designacion, observaciones, lugar, fecha_inicio,
            fecha_fin, perito_asignado, perito_id, desginacion,
            oficio_desplazamiento, estado
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        data.get('hoja_envio', ''),           # 1
        data.get('expediente', ''),           # 2
        data.get('dependencia', ''),          # 3
        data.get('tipo_perito', ''),          # 4
        data.get('tipo_actividad', ''),       # 5
        data.get('apoyos_tecnicos', ''),      # 6
        data.get('carpeta_fiscal', ''),       # 7
        data.get('denominacion', ''),         # 8 ← YA LO TIENES
        data.get('hoja_envio_designacion', ''), # 9
        data.get('observaciones', ''),        # 10
        data.get('lugar', ''),                # 11
        data['fecha_inicio'],                 # 12
        data['fecha_fin'],                    # 13
        data.get('perito_asignado', ''),      # 14
        data['perito_id'],                    # 15
        data.get('desginacion', ''),          # 16
        data.get('oficio_desplazamiento', ''), # 17
        'Pendiente'                           # 18 ← ESTADO (ESTE ES EL QUE FALTABA)
    ))
    
    asignacion_id = cursor.lastrowid
    conn.commit()
    conn.close()
    
    # Registrar en historial
    registrar_historial(asignacion_id, 'Creado', 'Asignación creada exitosamente')
    
    # Registrar en auditoría
    registrar_auditoria(
        session['usuario_id'],
        session.get('nombre_completo'),
        'CREAR_ASIGNACION',
        'ASIGNACIONES',
        f"Asignación creada: Expediente {data.get('expediente', 'S/N')}",
        asignacion_id
    )
    
    return jsonify({
        'success': True,
        'id': asignacion_id,
        'message': 'Asignación creada exitosamente'
    }), 201


@app.route('/api/asignacion/<int:id>', methods=['PUT'])
@login_required  # ← AGREGAR
def actualizar_asignacion(id):
    """
    Actualiza una asignación existente
    """
    data = request.json
    
    # Si se están cambiando las fechas, verificar disponibilidad
    if 'fecha_inicio' in data and 'fecha_fin' in data and 'perito_id' in data:
        disponible, conflictos = verificar_disponibilidad(
            data['perito_id'],
            data['fecha_inicio'],
            data['fecha_fin'],
            asignacion_id=id
        )
        
        if not disponible:
            return jsonify({
                'error': 'El perito no está disponible en estas fechas',
                'conflictos': conflictos
            }), 409
    
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    
    # Construir query de actualización dinámicamente
    campos = []
    valores = []
    
    campos_permitidos = [
        'hoja_envio', 'expediente', 'dependencia', 'tipo_perito', 'tipo_actividad', 'apoyo_tecnico',
        'carpeta_fiscal', 'hoja_envio_designacion','denominacion', 'observaciones', 'lugar', 'fecha_inicio',
        'fecha_fin', 'perito_asignado', 'perito_id', 'desginacion',
        'oficio_desplazamiento', 'estado'
    ]
    
    for campo in campos_permitidos:
        if campo in data:
            campos.append(f'{campo} = ?')
            valores.append(data[campo])
    
    if not campos:
        return jsonify({'error': 'No hay campos para actualizar'}), 400
    
    valores.append(id)
    query = f"UPDATE asignaciones SET {', '.join(campos)} WHERE id = ?"
    
    cursor.execute(query, valores)
    conn.commit()
    conn.close()
    
    # Registrar en historial
    registrar_historial(id, 'Modificado', f'Campos actualizados: {", ".join(campos)}')

    # ... código ...
    registrar_auditoria(
        session['usuario_id'],
        session.get('nombre_completo'),
        'EDITAR_ASIGNACION',
        'ASIGNACIONES',
        f"Asignación ID {id} editada",
        id
    )    
    
    return jsonify({
        'success': True,
        'message': 'Asignación actualizada exitosamente'
    })

# ------------------------------------------------------

@app.route('/api/asignacion/<int:id>', methods=['DELETE'])
@login_required  # ← AGREGAR
def eliminar_asignacion(id):
    """
    Elimina (marca como cancelada) una asignación
    """
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    
    # Verificar que la asignación existe
    cursor.execute('SELECT * FROM asignaciones WHERE id = ?', (id,))
    asignacion = cursor.fetchone()
    
    if not asignacion:
        conn.close()
        return jsonify({'error': 'Asignación no encontrada'}), 404
    
    # Marcar como cancelada en lugar de eliminar
    cursor.execute('UPDATE asignaciones SET estado = "Cancelado" WHERE id = ?', (id,))
    
    conn.commit()
    conn.close()
    
    # Registrar en historial
    registrar_historial(id, 'Cancelado', 'Asignación cancelada por el usuario')

    # ... código ...
    registrar_auditoria(
        session['usuario_id'],
        session.get('nombre_completo'),
        'ELIMINAR_ASIGNACION',
        'ASIGNACIONES',
        f"Asignación ID {id} eliminada",
        id
    )

    
    return jsonify({
        'success': True,
        'message': 'Asignación cancelada exitosamente'
    })


@app.route('/api/asignacion/<int:id>/cambiar-estado', methods=['PUT'])
@admin_required
def cambiar_estado(id):
    """
    Cambia el estado de una asignación
    """
    data = request.json
    nuevo_estado = data.get('estado')
    
    # Validar estado
    if nuevo_estado not in ['Pendiente', 'En Proceso', 'Completado', 'Cancelado']:
        return jsonify({'success': False, 'error': 'Estado inválido'}), 400
    
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    
    # Verificar que existe
    cursor.execute('SELECT id FROM asignaciones WHERE id = ?', (id,))
    if not cursor.fetchone():
        conn.close()
        return jsonify({'success': False, 'error': 'Asignación no encontrada'}), 404
    
    # Actualizar estado
    cursor.execute('UPDATE asignaciones SET estado = ? WHERE id = ?', (nuevo_estado, id))
    conn.commit()
    conn.close()
    
    # ← AGREGAR AUDITORÍA AQUÍ
    registrar_auditoria(
        session['usuario_id'],
        session.get('nombre_completo'),
        'CAMBIAR_ESTADO',
        'ASIGNACIONES',
        f"Estado cambiado a '{nuevo_estado}' en asignación ID {id}",
        id
    )
    
    return jsonify({'success': True, 'message': 'Estado actualizado correctamente'})

@app.route('/api/verificar-disponibilidad', methods=['POST'])
@login_required
def api_verificar_disponibilidad():
    """
    Verifica disponibilidad de un perito en tiempo real
    """
    data = request.json
    
    if not all(k in data for k in ('perito_id', 'fecha_inicio', 'fecha_fin')):
        return jsonify({'error': 'Faltan datos requeridos'}), 400
    
    disponible, conflictos = verificar_disponibilidad(
        data['perito_id'],
        data['fecha_inicio'],
        data['fecha_fin'],
        data.get('asignacion_id')
    )
    
    # Los conflictos ya vienen como diccionarios desde verificar_disponibilidad
    # NO necesitamos transformarlos, solo devolverlos
    return jsonify({
        'disponible': disponible,
        'conflictos': conflictos
    })

@app.route('/api/peritos', methods=['GET'])
@login_required
def get_peritos():
    """
    Obtiene lista de todos los peritos
    """
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    
    cursor.execute('SELECT * FROM peritos WHERE estado = "Activo" ORDER BY tipo, nombre_completo')
    
    peritos = []
    for row in cursor.fetchall():
        peritos.append({
            'id': row[0],
            'nombre': row[1],
            'tipo': row[2],
            'estado': row[3]
        })
    
    conn.close()
    return jsonify(peritos)

@app.route('/api/estadisticas', methods=['GET'])
@login_required
def get_estadisticas():
    """
    Obtiene estadísticas generales del sistema
    """
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    
    # Total de asignaciones por estado
    cursor.execute('''
        SELECT estado, COUNT(*) 
        FROM asignaciones 
        GROUP BY estado
    ''')
    por_estado = dict(cursor.fetchall())
    
    # Asignaciones por tipo de perito
    cursor.execute('''
        SELECT tipo_perito, COUNT(*) 
        FROM asignaciones 
        WHERE tipo_perito != ""
        GROUP BY tipo_perito
    ''')
    por_tipo = dict(cursor.fetchall())
    
    # Peritos más asignados
    cursor.execute('''
        SELECT p.nombre_completo, COUNT(a.id) as total
        FROM peritos p
        LEFT JOIN asignaciones a ON p.id = a.perito_id
        GROUP BY p.id
        ORDER BY total DESC
        LIMIT 5
    ''')
    top_peritos = [{'nombre': row[0], 'total': row[1]} for row in cursor.fetchall()]
    
    # Asignaciones por mes (últimos 6 meses)
    cursor.execute('''
        SELECT strftime('%Y-%m', fecha_inicio) as mes, COUNT(*) as total
        FROM asignaciones
        WHERE fecha_inicio >= date('now', '-6 months')
        GROUP BY mes
        ORDER BY mes
    ''')
    por_mes = [{'mes': row[0], 'total': row[1]} for row in cursor.fetchall()]
    
    conn.close()
    
    return jsonify({
        'por_estado': por_estado,
        'por_tipo': por_tipo,
        'top_peritos': top_peritos,
        'por_mes': por_mes
    })

@app.route('/api/buscar', methods=['GET'])
@login_required
def buscar_asignaciones():
    """
    Búsqueda avanzada de asignaciones
    Query params: q (término de búsqueda), campo (campo específico)
    """
    termino = request.args.get('q', '').strip()
    campo = request.args.get('campo', 'todos')
    
    if not termino:
        return jsonify([])
    
    conn = sqlite3.connect('database.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # Determinar si filtrar por perito
    perito_id = session.get('perito_id')
    es_admin = session.get('rol') == 'admin'
    
    # ... al inicio ...
    query = request.args.get('q', '')
    
    if query:
        registrar_auditoria(
            session['usuario_id'],
            session.get('nombre_completo'),
            'BUSCAR',
            'BUSQUEDAS',
            f"Búsqueda realizada: '{query}'",
            None
        )

    # Construir query según el campo
    if campo == 'todos':
        query = '''
            SELECT a.*, p.nombre_completo
            FROM asignaciones a
            LEFT JOIN peritos p ON a.perito_id = p.id
            WHERE a.estado != 'Cancelado'
            AND (
                a.hoja_envio LIKE ? 
                OR a.expediente LIKE ?
                OR a.carpeta_fiscal LIKE ?
                OR a.dependencia LIKE ?
                OR a.lugar LIKE ?
                OR a.observaciones LIKE ?
                OR p.nombre_completo LIKE ?
                OR a.perito_asignado LIKE ?
            )
        '''
        params = [f'%{termino}%'] * 8
    elif campo == 'expediente':
        query = '''
            SELECT a.*, p.nombre_completo
            FROM asignaciones a
            LEFT JOIN peritos p ON a.perito_id = p.id
            WHERE a.estado != 'Cancelado' AND a.expediente LIKE ?
        '''
        params = [f'%{termino}%']
    elif campo == 'carpeta_fiscal':
        query = '''
            SELECT a.*, p.nombre_completo
            FROM asignaciones a
            LEFT JOIN peritos p ON a.perito_id = p.id
            WHERE a.estado != 'Cancelado' AND a.carpeta_fiscal LIKE ?
        '''
        params = [f'%{termino}%']
    elif campo == 'dependencia':
        query = '''
            SELECT a.*, p.nombre_completo
            FROM asignaciones a
            LEFT JOIN peritos p ON a.perito_id = p.id
            WHERE a.estado != 'Cancelado' AND a.dependencia LIKE ?
        '''
        params = [f'%{termino}%']
    elif campo == 'lugar':
        query = '''
            SELECT a.*, p.nombre_completo
            FROM asignaciones a
            LEFT JOIN peritos p ON a.perito_id = p.id
            WHERE a.estado != 'Cancelado' AND a.lugar LIKE ?
        '''
        params = [f'%{termino}%']
    elif campo == 'perito':
        query = '''
            SELECT a.*, p.nombre_completo
            FROM asignaciones a
            LEFT JOIN peritos p ON a.perito_id = p.id
            WHERE a.estado != 'Cancelado' 
            AND (p.nombre_completo LIKE ? OR a.perito_asignado LIKE ?)
        '''
        params = [f'%{termino}%', f'%{termino}%']
    else:
        query = '''
            SELECT a.*, p.nombre_completo
            FROM asignaciones a
            LEFT JOIN peritos p ON a.perito_id = p.id
            WHERE a.estado != 'Cancelado'
        '''
        params = []
    
    # Filtrar por rol (perito solo ve las suyas)
    if not es_admin and perito_id:
        query += ' AND a.perito_id = ?'
        params.append(perito_id)
    
    query += ' ORDER BY a.fecha_registro DESC'
    
    cursor.execute(query, params)
    
    resultados = []
    for row in cursor.fetchall():
        resultados.append({
            'id': row['id'],
            'hoja_envio': row['hoja_envio'],
            'expediente': row['expediente'],
            'carpeta_fiscal': row['carpeta_fiscal'],
            'dependencia': row['dependencia'],
            'tipo_perito': row['tipo_perito'],
            'tipo_actividad': row['tipo_actividad'],
            'observaciones': row['observaciones'],
            'lugar': row['lugar'],
            'fecha_inicio': row['fecha_inicio'],
            'fecha_fin': row['fecha_fin'],
            'perito_asignado': row['perito_asignado'],
            'desginacion': row['desginacion'],
            'estado': row['estado'],
            'perito_nombre': row['nombre_completo'],
            'perito_id': row['perito_id']
        })
    
    conn.close()
    return jsonify(resultados)

# ============================================================================
# EXPORTACIÓN DE DATOS
# ============================================================================

@app.route('/api/exportar/excel', methods=['GET'])
@login_required
def exportar_excel():
    """
    Exporta asignaciones a formato Excel con formato profesional
    """
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    
    # Obtener filtros de la query string
    query = '''
        SELECT a.id, a.hoja_envio, a.expediente, a.dependencia, a.tipo_perito,
               a.tipo_actividad, a.apoyo_tecnico, a.carpeta_fiscal, a.hoja_envio_designacion, 
               a.observaciones, a.lugar, a.fecha_inicio, a.fecha_fin, 
               a.perito_asignado, a.desginacion, a.oficio_desplazamiento, 
               a.estado, a.fecha_registro
        FROM asignaciones a
        WHERE 1=1
    '''
    params = []
    
    if request.args.get('estado'):
        query += ' AND a.estado = ?'
        params.append(request.args.get('estado'))
    
    if request.args.get('fecha_desde'):
        query += ' AND a.fecha_inicio >= ?'
        params.append(request.args.get('fecha_desde'))
    
    if request.args.get('fecha_hasta'):
        query += ' AND a.fecha_fin <= ?'
        params.append(request.args.get('fecha_hasta'))
    
    query += ' ORDER BY a.fecha_inicio DESC'
    
    cursor.execute(query, params)
    datos = cursor.fetchall()
    conn.close()
    
    # Crear libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Asignaciones"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = [
        'ID', 'Hoja Envío', 'Expediente', 'Dependencia', 'Tipo Perito', 'Tipo Actividad', 'Apoyo Técnico',
        'Carpeta Fiscal', 'Hoja Envío Designación', 'Observaciones', 'Lugar', 'Fecha Inicio',
        'Fecha Fin', 'Perito Asignado', 'Designación', 'Oficio Desplazamiento',
        'Estado', 'Fecha Registro'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Datos
    for row_idx, row_data in enumerate(datos, 2):
        for col_idx, value in enumerate(row_data[:18], 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True)
            
            # Color según estado
            if col_idx == 14:  # Columna de estado
                if value == 'Completado':
                    cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                elif value == 'En Proceso':
                    cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                elif value == 'Pendiente':
                    cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
                elif value == 'Cancelado':
                    cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    
    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Guardar archivo
    filename = f'asignaciones_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    filepath = os.path.join('exports', filename)
    
    # Crear carpeta exports si no existe
    os.makedirs('exports', exist_ok=True)
    
    wb.save(filepath)
    
    return send_file(filepath, as_attachment=True, download_name=filename)

@app.route('/api/exportar/pdf', methods=['GET'])
@login_required
def exportar_pdf():
    """
    Exporta asignaciones a formato PDF con tabla profesional
    """
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    
    # Obtener datos (similar al Excel)
    query = '''
        SELECT a.hoja_envio, a.expediente, a.fecha_inicio, a.fecha_fin, 
               p.nombre_completo, a.estado, a.lugar
        FROM asignaciones a
        LEFT JOIN peritos p ON a.perito_id = p.id
        WHERE 1=1
    '''
    params = []
    
    if request.args.get('estado'):
        query += ' AND a.estado = ?'
        params.append(request.args.get('estado'))
    
    query += ' ORDER BY a.fecha_inicio DESC LIMIT 50'
    
    cursor.execute(query, params)
    datos = cursor.fetchall()
    conn.close()
    
    # Crear PDF
    filename = f'reporte_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    filepath = os.path.join('exports', filename)
    os.makedirs('exports', exist_ok=True)
    
    doc = SimpleDocTemplate(filepath, pagesize=A4)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    
    # Título
    title = Paragraph("<b>REPORTE DE ASIGNACIONES DE PERITOS</b>", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    # Fecha del reporte
    fecha_reporte = Paragraph(
        f"<b>Fecha:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        styles['Normal']
    )
    elements.append(fecha_reporte)
    elements.append(Spacer(1, 0.3*inch))
    
    # Tabla de datos
    data = [['Oficio', 'Expediente', 'F. Inicio', 'F. Fin', 'Perito', 'Estado', 'Lugar']]
    
    for row in datos:
        data.append([
            row[0][:15] if row[0] else '',
            row[1][:15] if row[1] else '',
            row[2],
            row[3],
            row[4][:25] if row[4] else '',
            row[5],
            row[6][:20] if row[6] else ''
        ])
    
    table = Table(data, colWidths=[1.2*inch, 1.2*inch, 0.9*inch, 0.9*inch, 1.8*inch, 0.9*inch, 1.5*inch])
    
    # Estilo de la tabla
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    elements.append(table)
    
    # Construir PDF
    doc.build(elements)
    
    return send_file(filepath, as_attachment=True, download_name=filename)


# ============================================
# ENDPOINTS PARA MANEJO DE ARCHIVOS
# ============================================


@app.route('/api/asignacion/<int:id>/subir-archivos', methods=['POST'])
@login_required
def subir_archivos(id):
    """
    Sube múltiples archivos adjuntos a una asignación
    """
    # Verificar permisos: admin puede subir a cualquiera, perito solo a las suyas
    es_admin = session.get('rol') == 'admin'
    perito_id = session.get('perito_id')
    
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    
    # Obtener asignación
    cursor.execute('SELECT archivos_adjuntos, perito_id FROM asignaciones WHERE id = ?', (id,))
    resultado = cursor.fetchone()
    
    
    if not resultado:
        conn.close()
        return jsonify({'error': 'Asignación no encontrada'}), 404
    
    # Verificar permisos (perito solo puede subir a sus asignaciones)
    if not es_admin and resultado[1] != perito_id:
        conn.close()
        return jsonify({'error': 'No tienes permiso para subir archivos a esta asignación'}), 403
    
    # Verificar que se enviaron archivos
    if 'archivos' not in request.files:
        conn.close()
        return jsonify({'error': 'No se enviaron archivos'}), 400
    
    archivos = request.files.getlist('archivos')
    archivos_guardados = []
    errores = []
    
    # Obtener archivos existentes
    archivos_existentes = []
    if resultado[0]:
        try:
            archivos_existentes = json.loads(resultado[0])
        except:
            archivos_existentes = []
    
    # Procesar cada archivo
    for archivo in archivos:
        if archivo.filename == '':
            continue
        
        if archivo and allowed_file(archivo.filename):
            filename = secure_filename(archivo.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')
            extension = filename.rsplit('.', 1)[1].lower()
            nombre_base = filename.rsplit('.', 1)[0][:30]
            filename_final = f"{id}_{timestamp}_{nombre_base}.{extension}"
            
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename_final)
            archivo.save(filepath)
            
            # Agregar información del usuario que subió el archivo
            archivos_guardados.append({
                'nombre_original': archivo.filename,
                'nombre_guardado': filename_final,
                'fecha_subida': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'tamano': os.path.getsize(filepath),
                'usuario_id': session['usuario_id'],  # ← NUEVO
                'usuario_nombre': session.get('nombre_completo', session['usuario'])  # ← NUEVO
            })
        else:
            errores.append(f"Archivo '{archivo.filename}' no permitido")
    
    # Combinar archivos nuevos con existentes
    todos_archivos = archivos_existentes + archivos_guardados
    
    # Guardar en base de datos
    cursor.execute(
        'UPDATE asignaciones SET archivos_adjuntos = ? WHERE id = ?',
        (json.dumps(todos_archivos), id)
    )
    conn.commit()
    conn.close()
    
    # Registrar en historial
    registrar_historial(
        id, 
        'Archivos Adjuntos', 
        f'Se agregaron {len(archivos_guardados)} archivo(s)'
    )
    
    # Registrar en auditoría
    nombres_archivos = [a['nombre_original'] for a in archivos_guardados]
    detalles = f"{len(archivos_guardados)} archivo(s) subido(s): {', '.join(nombres_archivos)}"

    registrar_auditoria(
        session['usuario_id'],
        session.get('nombre_completo'),
        'CARGA_ARCHIVO',
        'ASIGNACIONES',
        detalles,  # ← AQUÍ SE INCLUYEN LOS NOMBRES
        id
    )
    return jsonify({
        'success': True,
        'archivos_guardados': len(archivos_guardados),
        'archivos': archivos_guardados,
        'errores': errores
    })


@app.route('/api/asignacion/<int:id>/archivos', methods=['GET'])
@login_required  # ← AGREGAR
def listar_archivos(id):
    """
    Lista todos los archivos de una asignación
    """
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    
    cursor.execute('SELECT archivos_adjuntos FROM asignaciones WHERE id = ?', (id,))
    resultado = cursor.fetchone()
    conn.close()
    
    if not resultado:
        return jsonify({'error': 'Asignación no encontrada'}), 404
    
    archivos = []
    if resultado[0]:
        try:
            archivos = json.loads(resultado[0])
        except:
            archivos = []
    
    return jsonify({'archivos': archivos})



@app.route('/api/asignacion/<int:id>/archivo/<filename>', methods=['DELETE'])
@login_required
def eliminar_archivo(id, filename):
    """
    Elimina un archivo adjunto (Admin puede eliminar cualquiera, Perito solo los suyos)
    """
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    
    cursor.execute('SELECT archivos_adjuntos, perito_id FROM asignaciones WHERE id = ?', (id,))
    resultado = cursor.fetchone()
    
    if not resultado:
        conn.close()
        return jsonify({'error': 'Asignación no encontrada'}), 404
    
    archivos = []
    if resultado[0]:
        try:
            archivos = json.loads(resultado[0])
        except:
            archivos = []
    
    # Buscar el archivo a eliminar
    archivo_encontrado = None
    for a in archivos:
        if a['nombre_guardado'] == filename:
            archivo_encontrado = a
            break
    
    if not archivo_encontrado:
        conn.close()
        return jsonify({'error': 'Archivo no encontrado'}), 404
    
    # Verificar permisos
    es_admin = session.get('rol') == 'admin'
    usuario_id = session.get('usuario_id')
    perito_id = session.get('perito_id')
    
    # Admin puede eliminar cualquier archivo
    # Perito solo puede eliminar archivos que él subió Y de sus propias asignaciones
    if not es_admin:
        # Verificar que la asignación sea del perito
        if resultado[1] != perito_id:
            conn.close()
            return jsonify({'error': 'No tienes permiso para eliminar archivos de esta asignación'}), 403
        
        # Verificar que el archivo lo subió el perito
        if archivo_encontrado.get('usuario_id') != usuario_id:
            conn.close()
            return jsonify({'error': 'No tienes permiso para eliminar este archivo (no lo subiste tú)'}), 403
    
    # Filtrar el archivo a eliminar
    archivos_filtrados = [a for a in archivos if a['nombre_guardado'] != filename]
    
    # Eliminar archivo físico
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if os.path.exists(filepath):
        os.remove(filepath)
    
    # Actualizar base de datos
    cursor.execute(
        'UPDATE asignaciones SET archivos_adjuntos = ? WHERE id = ?',
        (json.dumps(archivos_filtrados), id)
    )
    conn.commit()
    conn.close()
    
    # Registrar en auditoría
    registrar_auditoria(
        session['usuario_id'],
        session.get('nombre_completo', session['usuario']),
        'ELIMINA_ARCHIVO',
        'ASIGNACIONES',
        f"Archivo '{archivo_encontrado['nombre_original']}' eliminado de asignación ID {id}",
        id
    )
    
    return jsonify({
        'success': True,
        'message': 'Archivo eliminado correctamente'
    })




@app.route('/uploads/<filename>')
@login_required  # ← AGREGAR
def descargar_archivo(filename):
    """
    Descarga un archivo adjunto
    """
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    
    if not os.path.exists(filepath):
        return jsonify({'error': 'Archivo no encontrado'}), 404
    
    return send_file(filepath, as_attachment=True)

# PARA ADNINB AACTIVIDADES

@app.route('/actividades-admin')
@admin_required
def actividades_admin():
    """
    Página para que admin vea todas las actividades de peritos
    """
    conn = sqlite3.connect('database.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT a.*, p.nombre_completo
        FROM actividades_peritos a
        LEFT JOIN peritos p ON a.perito_id = p.id
        ORDER BY a.fecha_inicio DESC
    ''')
    
    actividades = []
    from datetime import datetime, date
    hoy = date.today()
    
    for row in cursor.fetchall():
        fecha_inicio = datetime.strptime(row['fecha_inicio'], '%Y-%m-%d').date()
        fecha_fin = datetime.strptime(row['fecha_fin'], '%Y-%m-%d').date()
        
        # Calcular estado automático
        if hoy < fecha_inicio:
            estado_auto = 'Pendiente'
        elif fecha_inicio <= hoy <= fecha_fin:
            estado_auto = 'En Proceso'
        else:
            estado_auto = 'Completado'
        
        actividades.append({
            'id': row['id'],
            'perito_nombre': row['nombre_completo'],
            'tipo_perito': row['tipo_perito'],
            'dependencia': row['dependencia'],
            'carpeta_fiscal': row['carpeta_fiscal'],
            'denominacion': row['denominacion'],
            'tipo_actividad': row['tipo_actividad'],
            'fecha_inicio': row['fecha_inicio'],
            'fecha_fin': row['fecha_fin'],
            'hora_inicio': row['hora_inicio'],
            'hora_fin': row['hora_fin'],
            'estado': estado_auto
        })
    
    conn.close()
    
    return render_template('admin/actividades.html', actividades=actividades)



# ============================================
# ACTIVIDADES DE PERITOS
# ============================================

@app.route('/mis-actividades')
@login_required
def mis_actividades():
    """
    Página para que peritos registren sus actividades
    """
    # Solo peritos pueden acceder
    if session.get('rol') != 'perito':
        return redirect(url_for('index'))
    
    perito_id = session.get('perito_id')
    
    conn = sqlite3.connect('database.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # Obtener info del perito
    cursor.execute('SELECT nombre_completo, tipo FROM peritos WHERE id = ?', (perito_id,))
    perito = cursor.fetchone()
    
    # Obtener actividades del perito
    cursor.execute('''
        SELECT * FROM actividades_peritos 
        WHERE perito_id = ? 
        ORDER BY fecha_inicio DESC
    ''', (perito_id,))
    
    actividades = []
    from datetime import datetime, date
    hoy = date.today()
    
    for row in cursor.fetchall():
        fecha_inicio = datetime.strptime(row['fecha_inicio'], '%Y-%m-%d').date()
        fecha_fin = datetime.strptime(row['fecha_fin'], '%Y-%m-%d').date()
        
        # Calcular estado automático
        if hoy < fecha_inicio:
            estado_auto = 'Pendiente'
        elif fecha_inicio <= hoy <= fecha_fin:
            estado_auto = 'En Proceso'
        else:  # hoy > fecha_fin
            estado_auto = 'Completado'
        
        actividades.append({
            'id': row['id'],
            'dependencia': row['dependencia'],
            'carpeta_fiscal': row['carpeta_fiscal'],
            'denominacion': row['denominacion'],
            'tipo_actividad': row['tipo_actividad'],
            'apoyo_tecnico': row['apoyo_tecnico'],
            'observaciones': row['observaciones'],
            'fecha_inicio': row['fecha_inicio'],
            'fecha_fin': row['fecha_fin'],
            'hora_inicio': row['hora_inicio'],
            'hora_fin': row['hora_fin'],
            'fecha_registro': row['fecha_registro'],
            'estado': estado_auto  # Estado calculado automáticamente
        })
    
    conn.close()
    
    return render_template('core/mis_actividades.html',
                          perito_nombre=perito['nombre_completo'],
                          tipo_perito=perito['tipo'],
                          actividades=actividades)


@app.route('/api/actividad', methods=['POST'])
@login_required
def crear_actividad():
    """
    Crear nueva actividad de perito
    """
    # Solo peritos pueden crear actividades
    if session.get('rol') != 'perito':
        return jsonify({'success': False, 'error': 'No autorizado'}), 403
    
    data = request.json
    perito_id = session.get('perito_id')
    
    # Validaciones
    if not data.get('tipo_actividad') or not data.get('fecha_inicio') or not data.get('fecha_fin'):
        return jsonify({'success': False, 'error': 'Faltan campos requeridos'}), 400
    
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    
    # Obtener tipo de perito
    cursor.execute('SELECT tipo FROM peritos WHERE id = ?', (perito_id,))
    result = cursor.fetchone()
    tipo_perito = result[0] if result else ''
    
    try:
        cursor.execute('''
            INSERT INTO actividades_peritos (
                perito_id, tipo_perito, dependencia, carpeta_fiscal, denominacion,
                tipo_actividad, apoyo_tecnico, observaciones,
                fecha_inicio, fecha_fin, hora_inicio, hora_fin, estado
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            perito_id,
            tipo_perito,
            data.get('dependencia', ''),
            data.get('carpeta_fiscal', ''),
            data.get('denominacion', ''),
            data['tipo_actividad'],
            data.get('apoyo_tecnico', ''),
            data.get('observaciones', ''),
            data['fecha_inicio'],
            data['fecha_fin'],
            data.get('hora_inicio', ''),
            data.get('hora_fin', ''),
            'Pendiente'  # ← Estado por defecto
        ))
        
        actividad_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        # Registrar en auditoría
        registrar_auditoria(
            session['usuario_id'],
            session.get('nombre_completo'),
            'CREAR_ACTIVIDAD',
            'ACTIVIDADES',
            f"Actividad '{data['tipo_actividad']}' registrada",
            actividad_id
        )
        
        return jsonify({'success': True, 'id': actividad_id, 'message': 'Actividad registrada exitosamente'})
    
    except Exception as e:
        conn.close()
        print(f"Error al crear actividad: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/actividad/<int:id>', methods=['GET'])
@login_required
def get_actividad(id):
    """
    Obtener detalle de una actividad
    """
    conn = sqlite3.connect('database.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT a.*, p.nombre_completo
        FROM actividades_peritos a
        LEFT JOIN peritos p ON a.perito_id = p.id
        WHERE a.id = ?
    ''', (id,))
    
    row = cursor.fetchone()
    
    if not row:
        conn.close()
        return jsonify({'error': 'Actividad no encontrada'}), 404
    
    # Verificar permisos
    if session.get('rol') != 'admin' and row['perito_id'] != session.get('perito_id'):
        conn.close()
        return jsonify({'error': 'No autorizado'}), 403
    
    # Parsear archivos adjuntos si existen
    archivos = []
    if row['archivos_adjuntos']:
        try:
            archivos = json.loads(row['archivos_adjuntos'])
        except:
            archivos = []
    
    actividad = {
        'id': row['id'],
        'perito_id': row['perito_id'],
        'perito_nombre': row['nombre_completo'],
        'tipo_perito': row['tipo_perito'],
        'dependencia': row['dependencia'],
        'carpeta_fiscal': row['carpeta_fiscal'],
        'denominacion': row['denominacion'],
        'tipo_actividad': row['tipo_actividad'],
        'apoyo_tecnico': row['apoyo_tecnico'],
        'observaciones': row['observaciones'],
        'fecha_inicio': row['fecha_inicio'],
        'fecha_fin': row['fecha_fin'],
        'hora_inicio': row['hora_inicio'],
        'hora_fin': row['hora_fin'],
        'estado': row['estado'] if 'estado' in row.keys() else 'Pendiente',  # ← AGREGAR
        'archivos_adjuntos': archivos
    }
    
    conn.close()
    
    return jsonify(actividad)


@app.route('/api/actividad/<int:id>', methods=['DELETE'])
@login_required
def eliminar_actividad(id):
    """
    Eliminar actividad
    """
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    
    # Verificar que existe y pertenece al perito
    cursor.execute('SELECT perito_id FROM actividades_peritos WHERE id = ?', (id,))
    result = cursor.fetchone()
    
    if not result:
        conn.close()
        return jsonify({'success': False, 'error': 'Actividad no encontrada'}), 404
    
    # Verificar permisos
    if session.get('rol') != 'admin' and result[0] != session.get('perito_id'):
        conn.close()
        return jsonify({'success': False, 'error': 'No autorizado'}), 403
    
    cursor.execute('DELETE FROM actividades_peritos WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    
    # Registrar en auditoría
    registrar_auditoria(
        session['usuario_id'],
        session.get('nombre_completo'),
        'ELIMINAR_ACTIVIDAD',
        'ACTIVIDADES',
        f"Actividad ID {id} eliminada",
        id
    )
    
    return jsonify({'success': True, 'message': 'Actividad eliminada exitosamente'})

#--------------------------------------------
#CAMBIR DE ESTADO EN ACTIVIDAD
#--------------------------------------------
@app.route('/api/actividad/<int:id>/cambiar-estado', methods=['POST'])
@login_required
def cambiar_estado_actividad(id):
    """
    Cambia el estado de una actividad (solo perito dueño o admin)
    """
    data = request.json
    nuevo_estado = data.get('estado')
    
    if nuevo_estado not in ['Pendiente', 'En Proceso', 'Completado']:
        return jsonify({'success': False, 'error': 'Estado inválido'}), 400
    
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    
    # Verificar que existe y permisos
    cursor.execute('SELECT perito_id FROM actividades_peritos WHERE id = ?', (id,))
    result = cursor.fetchone()
    
    if not result:
        conn.close()
        return jsonify({'success': False, 'error': 'Actividad no encontrada'}), 404
    
    es_admin = session.get('rol') == 'admin'
    if not es_admin and result[0] != session.get('perito_id'):
        conn.close()
        return jsonify({'success': False, 'error': 'No autorizado'}), 403
    
    cursor.execute('UPDATE actividades_peritos SET estado = ? WHERE id = ?', (nuevo_estado, id))
    conn.commit()
    conn.close()
    
    # Auditoría
    registrar_auditoria(
        session['usuario_id'],
        session.get('nombre_completo'),
        'CAMBIAR_ESTADO_ACTIVIDAD',
        'ACTIVIDADES',
        f"Estado cambiado a '{nuevo_estado}' en actividad ID {id}",
        id
    )
    
    return jsonify({'success': True, 'message': 'Estado actualizado'})


#--------------------------------------------
#   EDITAR ACTIVIDAD ddddddddddddddddddddddddd
#--------------------------------------------
@app.route('/api/actividad/<int:id>', methods=['PUT'])
@login_required
def editar_actividad(id):
    """
    Edita una actividad (solo perito dueño o admin)
    """
    data = request.json
    
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    
    # Verificar permisos
    cursor.execute('SELECT perito_id FROM actividades_peritos WHERE id = ?', (id,))
    result = cursor.fetchone()
    
    if not result:
        conn.close()
        return jsonify({'success': False, 'error': 'Actividad no encontrada'}), 404
    
    es_admin = session.get('rol') == 'admin'
    if not es_admin and result[0] != session.get('perito_id'):
        conn.close()
        return jsonify({'success': False, 'error': 'No autorizado'}), 403
    
    cursor.execute('''
        UPDATE actividades_peritos
        SET dependencia = ?, carpeta_fiscal = ?, denominacion = ?,
            tipo_actividad = ?, apoyo_tecnico = ?, observaciones = ?,
            fecha_inicio = ?, fecha_fin = ?, hora_inicio = ?, hora_fin = ?
        WHERE id = ?
    ''', (
        data.get('dependencia', ''),
        data.get('carpeta_fiscal', ''),
        data.get('denominacion', ''),
        data['tipo_actividad'],
        data.get('apoyo_tecnico', ''),
        data.get('observaciones', ''),
        data['fecha_inicio'],
        data['fecha_fin'],
        data.get('hora_inicio', ''),
        data.get('hora_fin', ''),
        id
    ))
    
    conn.commit()
    conn.close()
    
    # Auditoría
    registrar_auditoria(
        session['usuario_id'],
        session.get('nombre_completo'),
        'EDITAR_ACTIVIDAD',
        'ACTIVIDADES',
        f"Actividad ID {id} editada",
        id
    )
    
    return jsonify({'success': True, 'message': 'Actividad actualizada'})


@app.route('/api/actividad/<int:id>/subir-archivos', methods=['POST'])
@login_required
def subir_archivos_actividad(id):
    """
    Sube archivos adjuntos a una actividad
    """
    # Verificar que la actividad existe y pertenece al perito
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    
    cursor.execute('SELECT archivos_adjuntos, perito_id FROM actividades_peritos WHERE id = ?', (id,))
    resultado = cursor.fetchone()
    
    if not resultado:
        conn.close()
        return jsonify({'error': 'Actividad no encontrada'}), 404
    
    # Verificar permisos
    es_admin = session.get('rol') == 'admin'
    perito_id = session.get('perito_id')
    
    if not es_admin and resultado[1] != perito_id:
        conn.close()
        return jsonify({'error': 'No tienes permiso para subir archivos a esta actividad'}), 403
    
    if 'archivos' not in request.files:
        conn.close()
        return jsonify({'error': 'No se enviaron archivos'}), 400
    
    archivos = request.files.getlist('archivos')
    archivos_guardados = []
    errores = []
    
    # Obtener archivos existentes
    archivos_existentes = []
    if resultado[0]:
        try:
            archivos_existentes = json.loads(resultado[0])
        except:
            archivos_existentes = []
    
    # Procesar cada archivo
    for archivo in archivos:
        if archivo.filename == '':
            continue
        
        if archivo and allowed_file(archivo.filename):
            filename = secure_filename(archivo.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')
            extension = filename.rsplit('.', 1)[1].lower()
            nombre_base = filename.rsplit('.', 1)[0][:30]
            filename_final = f"act_{id}_{timestamp}_{nombre_base}.{extension}"
            
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename_final)
            archivo.save(filepath)
            
            archivos_guardados.append({
                'nombre_original': archivo.filename,
                'nombre_guardado': filename_final,
                'fecha_subida': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'tamano': os.path.getsize(filepath),
                'usuario_id': session['usuario_id'],
                'usuario_nombre': session.get('nombre_completo', session['usuario'])
            })
        else:
            errores.append(f"Archivo '{archivo.filename}' no permitido")
    
    # Combinar archivos
    todos_archivos = archivos_existentes + archivos_guardados
    
    # Actualizar BD
    cursor.execute(
        'UPDATE actividades_peritos SET archivos_adjuntos = ? WHERE id = ?',
        (json.dumps(todos_archivos), id)
    )
    conn.commit()
    conn.close()
    
    # Auditoría
    # Registrar en auditoría
    nombres_archivos = [a['nombre_original'] for a in archivos_guardados]
    detalles = f"{len(archivos_guardados)} archivo(s) subido(s): {', '.join(nombres_archivos)}"

    registrar_auditoria(
        session['usuario_id'],
        session.get('nombre_completo'),
        'CARGA_ARCHIVO',
        'ACTIVIDADES',
        detalles,  # ← AQUÍ SE INCLUYEN LOS NOMBRES
        id
    )

    return jsonify({
        'success': True,
        'archivos_guardados': len(archivos_guardados),
        'archivos': archivos_guardados,
        'errores': errores
    })

@app.route('/api/actividad/<int:id>/archivo/<filename>', methods=['DELETE'])
@login_required
def eliminar_archivo_actividad(id, filename):
    """
    Elimina archivo de actividad (Admin elimina cualquiera, Perito solo los suyos)
    """
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    
    cursor.execute('SELECT archivos_adjuntos, perito_id FROM actividades_peritos WHERE id = ?', (id,))
    resultado = cursor.fetchone()
    
    if not resultado:
        conn.close()
        return jsonify({'error': 'Actividad no encontrada'}), 404
    
    archivos = []
    if resultado[0]:
        try:
            archivos = json.loads(resultado[0])
        except:
            archivos = []
    
    # Buscar el archivo
    archivo_encontrado = None
    for a in archivos:
        if a['nombre_guardado'] == filename:
            archivo_encontrado = a
            break
    
    if not archivo_encontrado:
        conn.close()
        return jsonify({'error': 'Archivo no encontrado'}), 404
    
    # Verificar permisos
    es_admin = session.get('rol') == 'admin'
    usuario_id = session.get('usuario_id')
    perito_id = session.get('perito_id')
    
    if not es_admin:
        if resultado[1] != perito_id:
            conn.close()
            return jsonify({'error': 'No tienes permiso'}), 403
        
        if archivo_encontrado.get('usuario_id') != usuario_id:
            conn.close()
            return jsonify({'error': 'No puedes eliminar este archivo (no lo subiste tú)'}), 403
    
    # Filtrar archivo
    archivos_filtrados = [a for a in archivos if a['nombre_guardado'] != filename]
    
    # Eliminar físicamente
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if os.path.exists(filepath):
        os.remove(filepath)
    
    # Actualizar BD
    cursor.execute(
        'UPDATE actividades_peritos SET archivos_adjuntos = ? WHERE id = ?',
        (json.dumps(archivos_filtrados), id)
    )
    conn.commit()
    conn.close()
    
    # Auditoría
    registrar_auditoria(
        session['usuario_id'],
        session.get('nombre_completo'),
        'ELIMINA_ARCHIVO',
        'ACTIVIDADES',
        f"Archivo '{archivo_encontrado['nombre_original']}' eliminado de actividad ID {id}",
        id
    )
    
    return jsonify({'success': True, 'message': 'Archivo eliminado correctamente'})

# ============================================
# ENDPOINTS PARA EDITAR
# ============================================




# ============================================
# ENDPOINTS PARA VACACIONES
# ============================================

@app.route('/vacaciones')
@login_required  # ← AGREGAR
def vacaciones():
    """
    Página de gestión de vacaciones
    """
    conn = sqlite3.connect('database.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # Obtener todas las vacaciones
    cursor.execute('''
        SELECT v.*, p.nombre_completo
        FROM vacaciones v
        LEFT JOIN peritos p ON v.perito_id = p.id
        ORDER BY v.fecha_inicio DESC
    ''')
    
    vacaciones_list = []
    for row in cursor.fetchall():
        vacaciones_list.append({
            'id': row['id'],
            'perito_id': row['perito_id'],
            'perito_nombre': row['nombre_completo'],
            'tipo_perito': row['tipo_perito'],
            'fecha_inicio': row['fecha_inicio'],
            'fecha_fin': row['fecha_fin'],
            'dias_totales': row['dias_totales'],
            'tipo_vacaciones': row['tipo_vacaciones'],
            'estado': row['estado'],
            'observaciones': row['observaciones'],
            'fecha_registro': row['fecha_registro']
        })
    
    # Estadísticas
    cursor.execute('SELECT COUNT(*) FROM vacaciones')
    total_vacaciones = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(*) FROM vacaciones WHERE estado = "Aprobadas"')
    aprobadas = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(*) FROM vacaciones WHERE estado = "Pendientes"')
    pendientes = cursor.fetchone()[0]
    
    conn.close()
    
    return render_template('admin/vacaciones.html',
                         vacaciones=vacaciones_list,
                         total=total_vacaciones,
                         aprobadas=aprobadas,
                         pendientes=pendientes)

@app.route('/api/vacacion', methods=['POST'])
@login_required  # ← AGREGAR
def crear_vacacion():
    """
    Registrar nuevo período de vacaciones
    """
    data = request.json
    
    # Validar datos requeridos
    if not all(k in data for k in ('perito_id', 'fecha_inicio', 'fecha_fin')):
        return jsonify({'error': 'Faltan datos requeridos'}), 400
    
    # Calcular días totales
    from datetime import datetime
    fecha_inicio = datetime.strptime(data['fecha_inicio'], '%Y-%m-%d')
    fecha_fin = datetime.strptime(data['fecha_fin'], '%Y-%m-%d')
    dias_totales = (fecha_fin - fecha_inicio).days + 1
    
    # Verificar si el perito ya tiene vacaciones en esas fechas
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT * FROM vacaciones
        WHERE perito_id = ?
        AND (
            (fecha_inicio <= ? AND fecha_fin >= ?) OR
            (fecha_inicio <= ? AND fecha_fin >= ?) OR
            (fecha_inicio >= ? AND fecha_fin <= ?)
        )
    ''', (data['perito_id'], data['fecha_fin'], data['fecha_inicio'],
          data['fecha_fin'], data['fecha_fin'], data['fecha_inicio'], data['fecha_fin']))
    
    conflictos = cursor.fetchall()
    
    if conflictos:
        conn.close()
        return jsonify({
            'error': 'El perito ya tiene vacaciones programadas en estas fechas',
            'conflictos': conflictos
        }), 409
    
    # Insertar vacaciones
    cursor.execute('''
        INSERT INTO vacaciones (
            perito_id, tipo_perito, fecha_inicio, fecha_fin,
            dias_totales, tipo_vacaciones, estado, observaciones
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        data['perito_id'],
        data.get('tipo_perito', ''),
        data['fecha_inicio'],
        data['fecha_fin'],
        dias_totales,
        data.get('tipo_vacaciones', 'Programadas'),
        data.get('estado', 'Aprobadas'),
        data.get('observaciones', '')
    ))
    
    vacacion_id = cursor.lastrowid
    conn.commit()
    conn.close()

    # ... al final, después del commit ...
    registrar_auditoria(
        session['usuario_id'],
        session.get('nombre_completo'),
        'CREAR_VACACION',
        'VACACIONES',
        f"Vacación registrada para perito ID {data['perito_id']}",
        vacacion_id
    )       
    
    return jsonify({
        'success': True,
        'id': vacacion_id,
        'dias_totales': dias_totales,
        'message': 'Vacaciones registradas exitosamente'
    }), 201

@app.route('/api/vacacion/<int:id>', methods=['GET'])
def get_vacacion(id):
    """
    Obtener detalles de vacaciones
    """
    conn = sqlite3.connect('database.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT v.*, p.nombre_completo
        FROM vacaciones v
        LEFT JOIN peritos p ON v.perito_id = p.id
        WHERE v.id = ?
    ''', (id,))
    
    row = cursor.fetchone()
    conn.close()
    
    if row:
        return jsonify({
            'id': row['id'],
            'perito_id': row['perito_id'],
            'perito_nombre': row['nombre_completo'],
            'tipo_perito': row['tipo_perito'],
            'fecha_inicio': row['fecha_inicio'],
            'fecha_fin': row['fecha_fin'],
            'dias_totales': row['dias_totales'],
            'tipo_vacaciones': row['tipo_vacaciones'],
            'estado': row['estado'],
            'observaciones': row['observaciones']
        })
    else:
        return jsonify({'error': 'Vacaciones no encontradas'}), 404

@app.route('/api/vacacion/<int:id>', methods=['DELETE'])
def eliminar_vacacion(id):
    """
    Eliminar registro de vacaciones
    """
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    
    cursor.execute('DELETE FROM vacaciones WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    # ... después del DELETE ...
    registrar_auditoria(
        session['usuario_id'],
        session.get('nombre_completo'),
        'ELIMINAR_VACACION',
        'VACACIONES',
        f"Vacación ID {id} eliminada",
        id
    )

    
    return jsonify({
        'success': True,
        'message': 'Vacaciones eliminadas correctamente'
    })

@app.route('/api/asignacion/<int:id>/imprimir-pdf')
@login_required
def imprimir_pdf_asignacion(id):
    """
    Genera un PDF con el detalle completo de una asignación
    """
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from io import BytesIO
    
    conn = sqlite3.connect('database.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # Obtener datos de la asignación
    cursor.execute('''
        SELECT a.*, p.nombre_completo as perito_nombre
        FROM asignaciones a
        LEFT JOIN peritos p ON a.perito_id = p.id
        WHERE a.id = ?
    ''', (id,))
    
    asignacion = cursor.fetchone()
    
    if not asignacion:
        conn.close()
        return jsonify({'error': 'Asignación no encontrada'}), 404
    
    # Verificar permisos
    if session.get('rol') != 'admin' and asignacion['perito_id'] != session.get('perito_id'):
        conn.close()
        return jsonify({'error': 'No autorizado'}), 403
    
    # Obtener archivos adjuntos
    archivos = []
    archivos_json = asignacion['archivos_adjuntos']
    
    print(f"🔍 DEBUG - Archivos JSON: {archivos_json}")  # DEBUG
    
    if archivos_json:
        try:
            archivos_list = json.loads(archivos_json)
            print(f"🔍 DEBUG - Archivos parseados: {archivos_list}")  # DEBUG
            
            for archivo in archivos_list:
                archivos.append({
                    'nombre_original': archivo.get('nombre_original', 'N/A'),
                    'tamano': archivo.get('tamano', 0),
                    'fecha_subida': archivo.get('fecha_subida', 'N/A'),
                    'usuario_nombre': archivo.get('usuario_nombre', '-')
                })
        except Exception as e:
            print(f"❌ Error al parsear archivos: {e}")
            archivos = []
    
    print(f"🔍 DEBUG - Total archivos: {len(archivos)}")  # DEBUG
    
    conn.close()
    
    # Crear PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=10,
        spaceBefore=15
    )
    
    normal_style = styles['Normal']
    
    # Contenido
    story = []
    
    # Logo (si existe)
    try:
        logo_path = 'static/imagenes/logompo.png'
        logo = Image(logo_path, width=1*inch, height=1*inch)
        story.append(logo)
        story.append(Spacer(1, 0.2*inch))
    except:
        pass
    
    # Título
    story.append(Paragraph("DETALLE DE ASIGNACIÓN", title_style))
    story.append(Paragraph(f"ID: #{asignacion['id']}", normal_style))
    story.append(Spacer(1, 0.3*inch))
    
    # Información General
    story.append(Paragraph("INFORMACIÓN GENERAL", heading_style))
    
    data_general = [
        ['Hoja de Envío:', asignacion['hoja_envio'] or '-'],
        ['Expediente:', asignacion['expediente'] or '-'],
        ['Dependencia:', asignacion['dependencia'] or '-'],
        ['Carpeta Fiscal:', asignacion['carpeta_fiscal'] or '-'],
        ['Denominación:', asignacion['denominacion'] or '-'],
        ['Tipo de Perito:', asignacion['tipo_perito'] or '-'],
        ['Tipo de Actividad:', asignacion['tipo_actividad'] or '-'],
    ]
    
    # Apoyo técnico
    apoyo = asignacion['apoyos_tecnicos'] or asignacion['apoyo_tecnico'] or '-'
    try:
        apoyos_list = json.loads(apoyo)
        apoyo = ', '.join(apoyos_list)
    except:
        pass
    data_general.append(['Apoyo Técnico:', apoyo])
    
    table_general = Table(data_general, colWidths=[2.5*inch, 4*inch])
    table_general.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e5e7eb')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    
    story.append(table_general)
    story.append(Spacer(1, 0.3*inch))
    
    # Perito Asignado
    story.append(Paragraph("PERITO ASIGNADO", heading_style))
    data_perito = [
        ['Nombre:', asignacion['perito_nombre'] or '-'],
    ]
    
    table_perito = Table(data_perito, colWidths=[2.5*inch, 4*inch])
    table_perito.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e5e7eb')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    
    story.append(table_perito)
    story.append(Spacer(1, 0.3*inch))
    
    # Fechas y Ubicación
    story.append(Paragraph("FECHAS Y UBICACIÓN", heading_style))
    data_fechas = [
        ['Fecha Inicio:', asignacion['fecha_inicio'] or '-'],
        ['Fecha Fin:', asignacion['fecha_fin'] or '-'],
        ['Lugar:', asignacion['lugar'] or '-'],
    ]
    
    table_fechas = Table(data_fechas, colWidths=[2.5*inch, 4*inch])
    table_fechas.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e5e7eb')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    
    story.append(table_fechas)
    story.append(Spacer(1, 0.3*inch))
    
    # Documentos Oficiales
    story.append(Paragraph("DOCUMENTOS OFICIALES", heading_style))
    data_docs = [
        ['Hoja Envío Designación:', asignacion['hoja_envio_designacion'] or '-'],
        ['Designación:', asignacion['desginacion'] or '-'],
        ['Oficio Desplazamiento:', asignacion['oficio_desplazamiento'] or '-'],
    ]
    
    table_docs = Table(data_docs, colWidths=[2.5*inch, 4*inch])
    table_docs.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e5e7eb')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    
    story.append(table_docs)
    story.append(Spacer(1, 0.3*inch))
    
    # Observaciones
    if asignacion['observaciones']:
        story.append(Paragraph("OBSERVACIONES", heading_style))
        story.append(Paragraph(asignacion['observaciones'], normal_style))
        story.append(Spacer(1, 0.3*inch))
    
    # Estado
    story.append(Paragraph("ESTADO", heading_style))
    story.append(Paragraph(f"<b>{asignacion['estado']}</b>", normal_style))
    story.append(Spacer(1, 0.3*inch))
    
    # Archivos Adjuntos
    # Archivos Adjuntos
    if archivos:
        story.append(Paragraph("ARCHIVOS ADJUNTOS", heading_style))
        
        data_archivos = [['Nombre', 'Tamaño', 'Fecha', 'Subido por']]
        for archivo in archivos:
            tamano = archivo.get('tamano', 0)
            tamano_mb = f"{(tamano / 1024 / 1024):.2f} MB" if tamano > 0 else 'N/A'
            data_archivos.append([
                archivo.get('nombre_original', 'N/A'),
                tamano_mb,
                archivo.get('fecha_subida', 'N/A'),
                archivo.get('usuario_nombre', '-')
            ])
        

        table_archivos = Table(data_archivos, colWidths=[2.5*inch, 1*inch, 1.5*inch, 1.5*inch])
        table_archivos.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        story.append(table_archivos)
    
    # Generar PDF
    doc.build(story)
    buffer.seek(0)
    
    # Registrar auditoría
    registrar_auditoria(
        session['usuario_id'],
        session.get('nombre_completo'),
        'IMPRIMIR_PDF',
        'ASIGNACIONES',
        f"PDF generado para asignación ID {id}",
        id
    )
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'asignacion_{id}_detalle.pdf'
    )


@app.route('/api/calendario/imprimir-pdf')
@login_required
def imprimir_calendario_pdf():
    """
    Genera PDF del calendario en UNA SOLA PÁGINA con texto justificado
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from io import BytesIO
    from datetime import datetime
    import calendar
    
    mes = int(request.args.get('mes', datetime.now().month))
    anio = int(request.args.get('anio', datetime.now().year))
    
    conn = sqlite3.connect('database.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    primer_dia = f"{anio}-{mes:02d}-01"
    ultimo_dia = f"{anio}-{mes:02d}-{calendar.monthrange(anio, mes)[1]}"
    
    # ← AGREGAR FILTRO DE CANCELADOS
    if session.get('rol') == 'admin' or session.get('rol') == 'invitado':
        cursor.execute('''
            SELECT a.*, p.nombre_completo as perito_nombre
            FROM asignaciones a
            LEFT JOIN peritos p ON a.perito_id = p.id
            WHERE fecha_inicio <= ? AND fecha_fin >= ? 
            AND a.estado != "Cancelado"
            ORDER BY fecha_inicio
        ''', (ultimo_dia, primer_dia))
    else:
        cursor.execute('''
            SELECT a.*, p.nombre_completo as perito_nombre
            FROM asignaciones a
            LEFT JOIN peritos p ON a.perito_id = p.id
            WHERE a.perito_id = ? AND fecha_inicio <= ? AND fecha_fin >= ?
            AND a.estado != "Cancelado"
            ORDER BY fecha_inicio
        ''', (session.get('perito_id'), ultimo_dia, primer_dia))
    
    asignaciones = cursor.fetchall()
    
    # Vacaciones (sin cambios)
    if session.get('rol') == 'admin' or session.get('rol') == 'invitado':
        cursor.execute('''
            SELECT v.*, p.nombre_completo as perito_nombre
            FROM vacaciones v
            LEFT JOIN peritos p ON v.perito_id = p.id
            WHERE fecha_inicio <= ? AND fecha_fin >= ?
        ''', (ultimo_dia, primer_dia))
    else:
        cursor.execute('''
            SELECT v.*, p.nombre_completo as perito_nombre
            FROM vacaciones v
            LEFT JOIN peritos p ON v.perito_id = p.id
            WHERE v.perito_id = ? AND fecha_inicio <= ? AND fecha_fin >= ?
        ''', (session.get('perito_id'), ultimo_dia, primer_dia))
    
    vacaciones = cursor.fetchall()
    
    # Actividades (sin cambios)
    if session.get('rol') == 'admin' or session.get('rol') == 'invitado':
        cursor.execute('''
            SELECT a.*, p.nombre_completo as perito_nombre
            FROM actividades_peritos a
            LEFT JOIN peritos p ON a.perito_id = p.id
            WHERE fecha_inicio <= ? AND fecha_fin >= ?
        ''', (ultimo_dia, primer_dia))
    else:
        cursor.execute('''
            SELECT a.*, p.nombre_completo as perito_nombre
            FROM actividades_peritos a
            LEFT JOIN peritos p ON a.perito_id = p.id
            WHERE a.perito_id = ? AND fecha_inicio <= ? AND fecha_fin >= ?
        ''', (session.get('perito_id'), ultimo_dia, primer_dia))
    
    actividades = cursor.fetchall()
    conn.close()
    
    # Crear PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(A4), 
        topMargin=0.25*inch, 
        bottomMargin=0.25*inch,
        leftMargin=0.25*inch, 
        rightMargin=0.25*inch
    )
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title',
        fontSize=14,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=6,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    # ← ESTILO CON WORDWRAP
    cell_style = ParagraphStyle(
        'Cell',
        fontSize=6,
        leading=7,
        alignment=TA_LEFT,
        wordWrap='CJK',
        leftIndent=0,
        rightIndent=0
    )
    
    story = []
    
    # Logo
    try:
        logo = Image('static/imagenes/logompo.png', width=0.4*inch, height=0.4*inch)
        story.append(logo)
        story.append(Spacer(1, 0.05*inch))
    except:
        pass
    
    # Título
    meses_es = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    
    titulo = f"{meses_es[mes].upper()} {anio}"
    if session.get('rol') != 'admin':
        titulo += f" - {session.get('nombre_completo')}"
    
    story.append(Paragraph(titulo, title_style))
    story.append(Spacer(1, 0.08*inch))
    
    # Calendario
    cal = calendar.monthcalendar(anio, mes)
    dias_semana = ['LUN', 'MAR', 'MIÉ', 'JUE', 'VIE', 'SÁB', 'DOM']
    data = [dias_semana]
    
    for semana in cal:
        fila = []
        for dia in semana:
            if dia == 0:
                fila.append('')
            else:
                fecha = f"{anio}-{mes:02d}-{dia:02d}"
                lineas = [f"<b><font size=8>{dia}</font></b>"]
                
                # Asignaciones con TODOS los datos
                for asig in asignaciones:
                    if asig['fecha_inicio'] <= fecha <= asig['fecha_fin']:
                        nombre = (asig['perito_nombre'] or 'N/A')[:16]
                        exp = (asig['expediente'] or 'S/E')[:10]
                        dep = (asig['dependencia'] or 'S/D')[:12]
                        
                        lineas.append(f"<font color='#1e40af'><b>📋 {nombre}</b></font>")
                        lineas.append(f"<font size=5>Exp: {exp}</font>")
                        lineas.append(f"<font size=5>Dep: {dep}</font>")
                        lineas.append("<font size=4>—</font>")
                
                # Actividades
                for act in actividades:
                    if act['fecha_inicio'] <= fecha <= act['fecha_fin']:
                        nombre = (act['perito_nombre'] or 'N/A')[:16]
                        tipo = (act['tipo_actividad'] or 'N/A')[:10]
                        
                        lineas.append(f"<font color='#7c3aed'><b>📌 {nombre}</b></font>")
                        lineas.append(f"<font size=5>{tipo}</font>")
                        lineas.append("<font size=4>—</font>")
                
                # Vacaciones
                for vac in vacaciones:
                    if vac['fecha_inicio'] <= fecha <= vac['fecha_fin']:
                        nombre = (vac['perito_nombre'] or 'N/A')[:16]
                        lineas.append(f"<font color='#16a34a'><b>🏖️ {nombre}</b></font>")
                        lineas.append("<font size=5>VACACIONES</font>")
                
                # Limpiar separador final
                if lineas[-1] == "<font size=4>—</font>":
                    lineas.pop()
                
                fila.append(Paragraph("<br/>".join(lineas), cell_style))
        
        data.append(fila)
    
    # Dimensiones ajustadas
    num_semanas = len(data) - 1
    col_width = 1.5*inch
    row_height_cell = 1.28*inch  # ← ALTURA FIJA para caber en una página
    
    row_heights = [0.22*inch] + [row_height_cell] * num_semanas
    
    table = Table(data, colWidths=[col_width]*7, rowHeights=row_heights)
    
    table.setStyle(TableStyle([
        # Cabecera
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
        ('TOPPADDING', (0, 0), (-1, 0), 5),
        
        # Celdas
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 1), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 1), (-1, -1), 2),
        ('RIGHTPADDING', (0, 1), (-1, -1), 2),
        ('TOPPADDING', (0, 1), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 2),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        
        # Fin de semana
        ('BACKGROUND', (5, 1), (5, -1), colors.HexColor('#f9f9f9')),
        ('BACKGROUND', (6, 1), (6, -1), colors.HexColor('#f9f9f9')),
    ]))
    
    story.append(table)
    story.append(Spacer(1, 0.08*inch))
    
    # Leyenda
    leyenda = ParagraphStyle('Leyenda', fontSize=7, alignment=TA_CENTER)
    story.append(Paragraph(
        "<b>LEYENDA:</b> <font color='#1e40af'>📋 Asignación</font> | "
        "<font color='#7c3aed'>📌 Actividad</font> | "
        "<font color='#16a34a'>🏖️ Vacación</font>", 
        leyenda
    ))
    
    doc.build(story)
    buffer.seek(0)
    
    # Auditoría
    registrar_auditoria(
        session['usuario_id'],
        session.get('nombre_completo'),
        'IMPRIMIR_CALENDARIO',
        'CALENDARIO',
        f"Calendario de {meses_es[mes]} {anio} impreso",
        None
    )
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'calendario_{meses_es[mes]}_{anio}.pdf'
    )

# ============================================================================
# #TODDO BIEN HASTA AQUI
# ============================================================================
# ============================================================================
# INICIALIZACIÓN Y EJECUCIÓN
# ============================================================================

if __name__ == '__main__':
    # Inicializar base de datos
    init_db()
    
    print("=" * 60)
    print("🚀 SISTEMA PERITO - Iniciado")
    print("=" * 60)
    print("📍 URL: http://127.0.0.1:4300")
    print("📊 Dashboard: http://127.0.0.1:4300")
    print("➕ Nueva Asignación: http://127.0.0.1:4300/nuevo")
    print("🔍 Búsqueda: http://127.0.0.1:4300/buscar")
    print("📅 Calendario: http://127.0.0.1:4300/calendario")
    print("👥 Peritos: http://127.0.0.1:4300/peritos")
    print("📈 Reportes: http://127.0.0.1:4300/reportes")
    print("=" * 60)
    print("💡 Presiona CTRL+C para detener el servidor")
    print("=" * 60)
    
    # Ejecutar aplicación en modo desarrollo
    app.run(host="0.0.0.0", port=4300, debug=True)